"""Pure tests for Phase 9A structured company-data intake."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest

from carbon_ledger import intake as intake_mod
from carbon_ledger.intake import (
    CONFIDENCE_HIGH,
    CONFIDENCE_LOW,
    CONFIDENCE_MEDIUM,
    MAX_UPLOAD_BYTES,
    ColumnMapping,
    IntakeError,
    IntakeMetadata,
    blank_template_csv_bytes,
    blank_template_xlsx_bytes,
    build_and_validate_intake,
    compute_bytes_sha256,
    default_value_maps,
    detect_header_row,
    is_reference_only_column,
    list_xlsx_sheet_names,
    parse_uploaded_table,
    parse_year_month_period,
    rank_xlsx_worksheets,
    reference_only_columns,
    sanitize_filename,
    source_document_id_from_hash,
    suggest_activity_type,
    suggest_column_mapping,
    suggest_column_mapping_with_confidence,
    suggest_unit,
    validate_upload_bytes,
    year_month_transform_preview,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
FIXED_INGESTED_AT = pd.Timestamp("2024-02-01T00:00:00Z")


def _metadata(**overrides: object) -> IntakeMetadata:
    base = IntakeMetadata(
        source_name="demo.csv",
        site_id="site_main",
        document_date=date(2024, 1, 31),
        data_quality_tier="unknown",
        intake_run_id="test_intake",
        ingested_at=FIXED_INGESTED_AT,
    )
    for key, value in overrides.items():
        setattr(base, key, value)
    return base


def _csv_bytes(text: str) -> bytes:
    return text.encode("utf-8")


def _valid_csv() -> bytes:
    return _csv_bytes(
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        "grid_electricity,50000,kWh,2024-01-01,2024-01-31\n"
        "natural_gas,8000,m3,2024-01-01,2024-01-31\n"
        "diesel,1200,L,2024-01-01,2024-01-31\n"
    )


def _mapping_for(table, **overrides: object) -> ColumnMapping:
    suggestions = suggest_column_mapping(list(table.columns))
    activity_map, unit_map = default_value_maps(
        table,
        ColumnMapping(
            activity_type_column=suggestions["activity_type"],
            activity_value_column=suggestions["activity_value"],
            unit_column=suggestions["unit"],
        ),
    )
    # Force complete mappings for known aliases.
    activity_map = {
        key: value or suggest_activity_type(key) for key, value in activity_map.items()
    }
    unit_map = {key: value or suggest_unit(key) for key, value in unit_map.items()}
    mapping = ColumnMapping(
        activity_type_column=suggestions["activity_type"],
        activity_value_column=suggestions["activity_value"],
        unit_column=suggestions["unit"],
        use_file_dates=True,
        start_date_column=suggestions["activity_start_date"],
        end_date_column=suggestions["activity_end_date"],
        activity_type_value_map=activity_map,
        unit_value_map=unit_map,
    )
    for key, value in overrides.items():
        setattr(mapping, key, value)
    return mapping


def _xlsx_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
    return buffer.getvalue()


def test_csv_bytes_parse_successfully() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    assert len(table.frame) == 3
    assert "activity_type" in table.columns


def test_utf8_bom_csv_parses() -> None:
    bom = b"\xef\xbb\xbf" + _valid_csv()
    table = parse_uploaded_table(file_name="bom.csv", data=bom)
    assert len(table.frame) == 3


def test_invalid_csv_encoding_rejected() -> None:
    bad = "活動,數量\n電力,1".encode("big5")
    with pytest.raises(IntakeError) as exc:
        parse_uploaded_table(file_name="bad.csv", data=bad)
    assert exc.value.code == "INVALID_ENCODING"


def test_xlsx_parses_successfully() -> None:
    data = _xlsx_bytes(
        {
            "Sheet1": pd.DataFrame(
                {
                    "activity_type": ["diesel"],
                    "activity_value": [10],
                    "unit": ["L"],
                    "activity_start_date": ["2024-01-01"],
                    "activity_end_date": ["2024-01-31"],
                }
            )
        }
    )
    table = parse_uploaded_table(file_name="demo.xlsx", data=data)
    assert len(table.frame) == 1
    assert table.sheet_name == "Sheet1"


def test_multiple_xlsx_sheets_discoverable() -> None:
    data = _xlsx_bytes(
        {
            "Alpha": pd.DataFrame({"a": [1]}),
            "Beta": pd.DataFrame({"b": [2]}),
        }
    )
    assert list_xlsx_sheet_names(data) == ["Alpha", "Beta"]


def test_selected_xlsx_sheet_is_parsed() -> None:
    data = _xlsx_bytes(
        {
            "Alpha": pd.DataFrame({"activity_type": ["diesel"], "x": [1]}),
            "Beta": pd.DataFrame({"activity_type": ["natural_gas"], "x": [2]}),
        }
    )
    table = parse_uploaded_table(
        file_name="multi.xlsx",
        data=data,
        sheet_name="Beta",
    )
    assert table.sheet_name == "Beta"
    assert table.frame.iloc[0]["activity_type"] == "natural_gas"


def test_unsupported_extension_rejected() -> None:
    with pytest.raises(IntakeError) as exc:
        validate_upload_bytes("scan.pdf", b"%PDF")
    assert exc.value.code == "UNSUPPORTED_FILE_TYPE"


def test_file_larger_than_10mb_rejected() -> None:
    with pytest.raises(IntakeError) as exc:
        validate_upload_bytes("big.csv", b"a" * (MAX_UPLOAD_BYTES + 1))
    assert exc.value.code == "FILE_TOO_LARGE"


def test_sanitized_filename_removes_path_components() -> None:
    assert sanitize_filename(r"C:\temp\..\secret\demo.csv") == "demo.csv"
    assert sanitize_filename("/var/tmp/upload.xlsx") == "upload.xlsx"


def test_sha256_is_deterministic() -> None:
    data = _valid_csv()
    assert compute_bytes_sha256(data) == compute_bytes_sha256(data)


def test_identical_bytes_produce_identical_source_document_id() -> None:
    data = _valid_csv()
    digest = compute_bytes_sha256(data)
    assert source_document_id_from_hash(digest) == source_document_id_from_hash(
        digest
    )


def test_identical_input_produces_identical_activity_record_ids() -> None:
    data = _valid_csv()
    table = parse_uploaded_table(file_name="demo.csv", data=data)
    mapping = _mapping_for(table)
    first = build_and_validate_intake(table, mapping, _metadata())
    second = build_and_validate_intake(table, mapping, _metadata())
    assert list(first.accepted_activities["record_id"]) == list(
        second.accepted_activities["record_id"]
    )


def test_column_alias_suggests_activity_type() -> None:
    assert suggest_column_mapping(["activity_type", "x"])["activity_type"] == (
        "activity_type"
    )


def test_chinese_alias_suggests_activity_type() -> None:
    assert suggest_column_mapping(["活動類型", "數量"])["activity_type"] == "活動類型"


def test_amount_alias_suggestion_works() -> None:
    assert suggest_column_mapping(["用量", "單位"])["activity_value"] == "用量"


def test_unit_alias_suggestion_works() -> None:
    assert suggest_column_mapping(["單位"])["unit"] == "單位"


def test_date_alias_suggestion_works() -> None:
    mapping = suggest_column_mapping(["開始日期", "結束日期"])
    assert mapping["activity_start_date"] == "開始日期"
    assert mapping["activity_end_date"] == "結束日期"


def test_unmatched_column_remains_unselected() -> None:
    assert suggest_column_mapping(["foo", "bar"])["activity_type"] == ""


def test_electricity_alias_maps_to_grid_electricity() -> None:
    assert suggest_activity_type("電力") == "grid_electricity"
    assert suggest_activity_type("electricity") == "grid_electricity"


def test_natural_gas_alias_suggestion_works() -> None:
    assert suggest_activity_type("天然氣") == "natural_gas"


def test_diesel_alias_suggestion_works() -> None:
    assert suggest_activity_type("柴油") == "diesel"


def test_steel_alias_suggestion_works() -> None:
    assert suggest_activity_type("採購鋼材") == "purchased_steel"
    assert suggest_activity_type("盤元") == "purchased_steel"


def test_production_alias_suggestion_works() -> None:
    assert suggest_activity_type("生產數量") == "finished_goods_output"


def test_unit_degree_suggests_kwh() -> None:
    assert suggest_unit("度") == "kWh"


def test_m3_symbol_suggests_m3() -> None:
    assert suggest_unit("m³") == "m3"


def test_liter_chinese_suggests_l() -> None:
    assert suggest_unit("公升") == "L"


def test_tonne_chinese_suggests_t() -> None:
    assert suggest_unit("噸") == "t"


def test_unmatched_activity_value_requires_mapping() -> None:
    assert suggest_activity_type("神秘燃料") == ""


def test_unmatched_unit_requires_mapping() -> None:
    assert suggest_unit("gallon") == ""


def _result_with_value(value: object) -> intake_mod.IntakeValidationResult:
    csv = (
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        f"diesel,{value},L,2024-01-01,2024-01-31\n"
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_csv_bytes(csv))
    return build_and_validate_intake(table, _mapping_for(table), _metadata())


def test_invalid_activity_amount_rejected() -> None:
    result = _result_with_value("abc")
    assert result.rejected_count == 1
    assert result.rejected_rows.iloc[0]["issue_code"] == "INVALID_ACTIVITY_VALUE"


def test_zero_activity_amount_rejected() -> None:
    result = _result_with_value(0)
    assert result.rejected_count == 1


def test_negative_activity_amount_rejected() -> None:
    result = _result_with_value(-5)
    assert result.rejected_count == 1


def test_invalid_date_rejected() -> None:
    csv = (
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        "diesel,10,L,not-a-date,2024-01-31\n"
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_csv_bytes(csv))
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.rejected_count == 1
    assert result.rejected_rows.iloc[0]["issue_code"] == "INVALID_DATE"


def test_end_date_before_start_date_rejected() -> None:
    csv = (
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        "diesel,10,L,2024-02-01,2024-01-01\n"
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_csv_bytes(csv))
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.rejected_count == 1
    assert result.rejected_rows.iloc[0]["issue_code"] == "INVALID_DATE"


def test_electricity_with_l_rejected_as_unit_mismatch() -> None:
    csv = (
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        "grid_electricity,10,L,2024-01-01,2024-01-31\n"
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_csv_bytes(csv))
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.rejected_count == 1
    assert result.rejected_rows.iloc[0]["issue_code"] == "ACTIVITY_UNIT_MISMATCH"


def test_gas_with_kwh_rejected_as_unit_mismatch() -> None:
    csv = (
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        "natural_gas,10,kWh,2024-01-01,2024-01-31\n"
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_csv_bytes(csv))
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.rejected_count == 1
    assert result.rejected_rows.iloc[0]["issue_code"] == "ACTIVITY_UNIT_MISMATCH"


def test_diesel_with_l_accepted_structurally() -> None:
    result = _result_with_value(12)
    assert result.accepted_count == 1
    assert result.rejected_count == 0


def test_purchased_steel_with_t_accepted_structurally() -> None:
    csv = (
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        "purchased_steel,15,t,2024-01-01,2024-01-31\n"
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_csv_bytes(csv))
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.accepted_count == 1


def test_source_document_uses_company_provided() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.source_documents.iloc[0]["data_origin"] == "company_provided"


def test_source_document_is_synthetic_false() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert bool(result.source_documents.iloc[0]["is_synthetic"]) is False


def test_source_document_hash_matches_uploaded_bytes() -> None:
    data = _valid_csv()
    table = parse_uploaded_table(file_name="demo.csv", data=data)
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.source_documents.iloc[0]["sha256"] == compute_bytes_sha256(data)


def test_document_type_other_includes_required_notes() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    row = result.source_documents.iloc[0]
    assert row["document_type"] == "other"
    assert str(row["notes"]).strip()


def test_row_locator_is_deterministic() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.accepted_activities.iloc[0]["source_locator"] == "row:2"


def test_unknown_conservative_fields_trigger_needs_review() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert (
        result.accepted_activities["human_review_status"] == "needs_review"
    ).all()


def test_valid_rows_remain_accepted_when_another_row_fails() -> None:
    csv = (
        "activity_type,activity_value,unit,activity_start_date,activity_end_date\n"
        "diesel,10,L,2024-01-01,2024-01-31\n"
        "diesel,-1,L,2024-01-01,2024-01-31\n"
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_csv_bytes(csv))
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.accepted_count == 1
    assert result.rejected_count == 1


def test_rejected_rows_contain_source_row_issue_and_reason() -> None:
    result = _result_with_value("abc")
    row = result.rejected_rows.iloc[0]
    assert "source_row" in row
    assert row["issue_code"]
    assert row["issue_message"]


def test_intake_does_not_mutate_uploaded_dataframe() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    before = table.frame.copy()
    build_and_validate_intake(table, _mapping_for(table), _metadata())
    pd.testing.assert_frame_equal(before, table.frame)


def test_intake_does_not_write_files_to_repository(tmp_path: Path) -> None:
    before = {path for path in REPO_ROOT.rglob("*") if path.is_file()}
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    build_and_validate_intake(table, _mapping_for(table), _metadata())
    after = {path for path in REPO_ROOT.rglob("*") if path.is_file()}
    assert before == after
    assert blank_template_csv_bytes().startswith(b"activity_type,")


def test_xlsx_template_has_three_sheets_and_is_not_auto_imported() -> None:
    from openpyxl import load_workbook

    payload = blank_template_xlsx_bytes()
    workbook = load_workbook(BytesIO(payload))
    assert workbook.sheetnames == ["資料填寫", "填寫範例", "欄位說明"]
    fill = workbook["資料填寫"]
    assert [cell.value for cell in fill[1]] == [
        "活動類型",
        "用量",
        "單位",
        "開始日期",
        "結束日期",
    ]
    assert fill["A2"].value is None
    example = workbook["填寫範例"]
    assert example["A2"].value == "外購電力"
    guide = workbook["欄位說明"]
    assert guide["D2"].value == "activity_type"
    assert blank_template_xlsx_bytes() == payload



def _real_world_workbook_bytes(*, data_sheet_name: str = "活動數據") -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    prose = workbook.active
    prose.title = "說明"
    prose["A1"] = "填寫說明"
    prose["A2"] = "本表僅供說明，請閱讀公司碳盤查作業指引後再填寫。"
    prose["A3"] = "注意：請勿修改排放係數計算方式說明文字。"

    data = workbook.create_sheet(data_sheet_name)
    data.append(
        ["廠區", "年月", "能源別", "使用量", "單位", "排放係數", "排放量 (kgCO2e)"]
    )
    data.append(["台北總部", "2025-01", "外購電力", 120000, "kWh", 0.494, 59280])
    data.append(["台北總部", "2025-02", "天然氣", 8000, "m3", 2.0, 16000])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _workbook_with_header_below_row1() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet2"
    sheet.append(["公司能源使用月報"])
    sheet.append(["本列是說明文字，不是欄位名稱"])
    sheet.append(["廠區", "年月", "能源別", "使用量", "單位"])
    sheet.append(["高雄廠", "2024-03", "柴油", 100, "L"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_worksheet_selection_does_not_depend_solely_on_sheet_name() -> None:
    data = _real_world_workbook_bytes(data_sheet_name="abc123")
    ranked = rank_xlsx_worksheets(data)
    assert ranked[0].sheet_name == "abc123"
    assert ranked[0].sheet_name != "說明"


def test_prose_instruction_sheet_not_preferred_over_tabular_data() -> None:
    data = _real_world_workbook_bytes(data_sheet_name="活動數據")
    ranked = rank_xlsx_worksheets(data)
    assert ranked[0].sheet_name == "活動數據"
    prose = next(item for item in ranked if item.sheet_name == "說明")
    assert ranked[0].score > prose.score


def test_arbitrary_data_sheet_name_still_ranks_correctly() -> None:
    data = _real_world_workbook_bytes(data_sheet_name="Sheet2")
    ranked = rank_xlsx_worksheets(data)
    assert ranked[0].sheet_name == "Sheet2"
    table = parse_uploaded_table(file_name="company.xlsx", data=data)
    assert table.sheet_name == "Sheet2"
    assert "能源別" in table.columns


def test_header_row_detected_below_row_1() -> None:
    data = _workbook_with_header_below_row1()
    raw = pd.read_excel(BytesIO(data), header=None, dtype=object, engine="openpyxl")
    detection = detect_header_row(raw)
    assert detection.header_row_index == 2
    table = parse_uploaded_table(file_name="offset.xlsx", data=data, header_row=2)
    assert list(table.columns)[:5] == ["廠區", "年月", "能源別", "使用量", "單位"]
    assert table.frame.iloc[0]["能源別"] == "柴油"


def test_energy_bie_suggests_activity_type() -> None:
    detailed = suggest_column_mapping_with_confidence(
        ["廠區", "年月", "能源別", "使用量", "單位"]
    )
    assert detailed["activity_type"].source_column == "能源別"
    assert detailed["activity_type"].confidence == CONFIDENCE_HIGH
    assert suggest_column_mapping(["能源別"])["activity_type"] == "能源別"


def test_usage_amount_suggests_activity_value() -> None:
    assert suggest_column_mapping(["使用量"])["activity_value"] == "使用量"
    assert (
        suggest_column_mapping_with_confidence(["使用量"])["activity_value"].confidence
        == CONFIDENCE_HIGH
    )


def test_plant_area_suggests_site() -> None:
    detailed = suggest_column_mapping_with_confidence(["廠區", "使用量"])
    assert detailed["site_id"].source_column == "廠區"
    assert detailed["site_id"].confidence == CONFIDENCE_HIGH


def test_year_month_recognized_as_monthly_period() -> None:
    detailed = suggest_column_mapping_with_confidence(
        ["廠區", "年月", "能源別", "使用量", "單位"]
    )
    assert detailed["year_month"].source_column == "年月"
    assert detailed["year_month"].confidence == CONFIDENCE_HIGH


def test_year_month_2025_01_converts_after_confirmation() -> None:
    preview = year_month_transform_preview("2025-01")
    assert preview["activity_start_date"] == "2025-01-01"
    assert preview["activity_end_date"] == "2025-01-31"
    start, end = parse_year_month_period("2025/01")
    assert start == date(2025, 1, 1)
    assert end == date(2025, 1, 31)
    start, end = parse_year_month_period("2025年1月")
    assert start == date(2025, 1, 1)
    assert end == date(2025, 1, 31)


def test_february_leap_year_logic() -> None:
    start, end = parse_year_month_period("2024-02")
    assert start == date(2024, 2, 1)
    assert end == date(2024, 2, 29)
    start, end = parse_year_month_period("2025-02")
    assert start == date(2025, 2, 1)
    assert end == date(2025, 2, 28)


def test_uploaded_emission_factor_not_used_as_registry() -> None:
    columns = ["能源別", "使用量", "單位", "排放係數", "排放量 (kgCO2e)"]
    assert is_reference_only_column("排放係數")
    assert is_reference_only_column("排放量 (kgCO2e)")
    assert set(reference_only_columns(columns)) >= {"排放係數", "排放量 (kgCO2e)"}
    suggestions = suggest_column_mapping(columns)
    assert suggestions["activity_value"] == "使用量"
    assert suggestions["activity_value"] != "排放量 (kgCO2e)"
    assert "排放係數" not in suggestions.values()


def test_uploaded_emission_amount_not_treated_as_calculated_truth() -> None:
    data = _real_world_workbook_bytes()
    table = parse_uploaded_table(
        file_name="company.xlsx",
        data=data,
        sheet_name="活動數據",
    )
    mapping = ColumnMapping(
        activity_type_column="能源別",
        activity_value_column="使用量",
        unit_column="單位",
        site_column="廠區",
        use_file_dates=False,
        use_year_month=True,
        year_month_column="年月",
        year_month_confirmed=True,
        activity_type_value_map={
            "外購電力": "grid_electricity",
            "天然氣": "natural_gas",
        },
        unit_value_map={"kWh": "kWh", "m3": "m3"},
    )
    result = build_and_validate_intake(table, mapping, _metadata())
    assert result.accepted_count == 2
    first = result.accepted_activities.iloc[0]
    assert float(first["activity_value"]) == 120000.0
    assert float(first["activity_value"]) != 59280.0
    assert "emission_factor" not in result.accepted_activities.columns
    assert "co2e" not in result.accepted_activities.columns
    assert first["activity_start_date"].date() == date(2025, 1, 1)
    assert first["activity_end_date"].date() == date(2025, 1, 31)
    assert first["site_id"] == "台北總部"


def test_year_month_requires_confirmation() -> None:
    data = _real_world_workbook_bytes()
    table = parse_uploaded_table(
        file_name="company.xlsx",
        data=data,
        sheet_name="活動數據",
    )
    mapping = ColumnMapping(
        activity_type_column="能源別",
        activity_value_column="使用量",
        unit_column="單位",
        use_year_month=True,
        year_month_column="年月",
        year_month_confirmed=False,
        activity_type_value_map={"外購電力": "grid_electricity"},
        unit_value_map={"kWh": "kWh"},
    )
    with pytest.raises(IntakeError) as exc:
        build_and_validate_intake(table, mapping, _metadata())
    assert exc.value.code == "YEAR_MONTH_NOT_CONFIRMED"


def test_user_can_override_every_suggestion() -> None:
    columns = ["能源別", "使用量", "單位", "備用類型", "備用數量"]
    suggestions = suggest_column_mapping(columns)
    assert suggestions["activity_type"] == "能源別"
    overridden = ColumnMapping(
        activity_type_column="備用類型",
        activity_value_column="備用數量",
        unit_column="單位",
        use_file_dates=False,
        period_start=date(2025, 1, 1),
        period_end=date(2025, 1, 31),
        activity_type_value_map={"外購電力": "grid_electricity"},
        unit_value_map={"kWh": "kWh"},
    )
    frame = pd.DataFrame(
        [
            {
                "能源別": "柴油",
                "使用量": 1,
                "單位": "kWh",
                "備用類型": "外購電力",
                "備用數量": 50,
            }
        ]
    )
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    table = intake_mod.UploadedTable(
        file_name=table.file_name,
        file_extension=table.file_extension,
        sha256=table.sha256,
        sheet_name=None,
        sheet_names=(),
        columns=tuple(frame.columns),
        frame=frame,
        byte_length=table.byte_length,
        header_row_index=0,
    )
    result = build_and_validate_intake(table, overridden, _metadata())
    assert result.accepted_count == 1
    assert result.accepted_activities.iloc[0]["activity_type"] == "grid_electricity"
    assert float(result.accepted_activities.iloc[0]["activity_value"]) == 50.0


def test_low_confidence_mappings_remain_unconfirmed() -> None:
    detailed = suggest_column_mapping_with_confidence(["foo", "bar", "baz"])
    assert detailed["activity_type"].confidence == CONFIDENCE_LOW
    assert detailed["activity_type"].source_column == ""
    assert suggest_column_mapping(["foo", "bar"])["activity_type"] == ""
    # Medium-confidence labels still suggest but are not auto-finalized by intake.
    medium = suggest_column_mapping_with_confidence(["項目", "數量", "單位"])
    assert medium["activity_type"].confidence == CONFIDENCE_MEDIUM
    assert medium["activity_type"].source_column == "項目"


def test_existing_canonical_template_uploads_still_work() -> None:
    table = parse_uploaded_table(file_name="demo.csv", data=_valid_csv())
    result = build_and_validate_intake(table, _mapping_for(table), _metadata())
    assert result.accepted_count == 3
    assert result.rejected_count == 0
    assert list(table.columns)[:5] == [
        "activity_type",
        "activity_value",
        "unit",
        "activity_start_date",
        "activity_end_date",
    ]


def test_real_world_workbook_end_to_end_suggestions() -> None:
    data = _real_world_workbook_bytes()
    ranked = rank_xlsx_worksheets(data)
    assert ranked[0].sheet_name == "活動數據"
    table = parse_uploaded_table(
        file_name="company.xlsx",
        data=data,
        sheet_name=ranked[0].sheet_name,
    )
    suggestions = suggest_column_mapping_with_confidence(list(table.columns))
    assert suggestions["activity_type"].source_column == "能源別"
    assert suggestions["activity_value"].source_column == "使用量"
    assert suggestions["unit"].source_column == "單位"
    assert suggestions["site_id"].source_column == "廠區"
    assert suggestions["year_month"].source_column == "年月"
    assert "排放係數" not in {
        item.source_column for item in suggestions.values() if item.source_column
    }
