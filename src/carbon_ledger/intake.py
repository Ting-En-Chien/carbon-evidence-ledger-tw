"""Phase 9A structured company-data intake (upload → map → validate).

Pure presentation-independent logic. Streamlit must not be imported here.
Uploaded bytes are processed in memory only; nothing is written to disk.

Interpretation helpers (worksheet ranking, header detection, column aliases,
year-month transforms) adapt to real-world spreadsheets. Users must confirm
uncertain mappings before canonical records are built.
"""

from __future__ import annotations

import calendar
import hashlib
import math
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from pandera.errors import SchemaError, SchemaErrors

from carbon_ledger.domain import (
    ACTIVITY_TYPE_UNIT_COMPATIBILITY,
    ACTIVITY_TYPES,
    SUPPORTED_UNITS,
)
from carbon_ledger.potential_duplicates import find_potential_duplicate_groups
from carbon_ledger.schemas import (
    validate_activity_records,
    validate_source_documents,
)

MAX_UPLOAD_MB = 10
MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024
SUPPORTED_EXTENSIONS = frozenset({".csv", ".xlsx"})
BLANK_TEMPLATE_COLUMNS = (
    "activity_type",
    "activity_value",
    "unit",
    "activity_start_date",
    "activity_end_date",
)
SOURCE_DOCUMENT_NOTES = "User-uploaded structured activity-data file."
UNMAPPED_SENTINEL = ""

CONFIDENCE_HIGH = "high"
CONFIDENCE_MEDIUM = "medium"
CONFIDENCE_LOW = "low"

ISSUE_MISSING_REQUIRED_MAPPING = "MISSING_REQUIRED_MAPPING"
ISSUE_INVALID_ACTIVITY_VALUE = "INVALID_ACTIVITY_VALUE"
ISSUE_UNSUPPORTED_UNIT = "UNSUPPORTED_UNIT"
ISSUE_ACTIVITY_UNIT_MISMATCH = "ACTIVITY_UNIT_MISMATCH"
ISSUE_INVALID_DATE = "INVALID_DATE"
ISSUE_UNMAPPED_ACTIVITY_TYPE = "UNMAPPED_ACTIVITY_TYPE"
ISSUE_UNMAPPED_UNIT = "UNMAPPED_UNIT"
ISSUE_SCHEMA_VALIDATION_FAILED = "SCHEMA_VALIDATION_FAILED"
ISSUE_UNSUPPORTED_FILE_TYPE = "UNSUPPORTED_FILE_TYPE"
ISSUE_FILE_TOO_LARGE = "FILE_TOO_LARGE"
ISSUE_INVALID_ENCODING = "INVALID_ENCODING"
ISSUE_YEAR_MONTH_NOT_CONFIRMED = "YEAR_MONTH_NOT_CONFIRMED"

# V1 electricity enterprise-inventory uses. Do not guess when unknown.
_ENTERPRISE_ELECTRICITY_USES = frozenset(
    {"general_factory", "heat_treatment", "forging"}
)
_NG_SUBTYPE_RE = re.compile(r"NG\s*([12])", re.IGNORECASE)
_DIESEL_VEHICLE_HINTS = (
    "公司車輛",
    "公務車",
    "company vehicle",
    "company-vehicle",
    "fleet diesel",
)
READINESS_READY = "ready"
READINESS_NEEDS_CONFIRM = "needs_confirm"
READINESS_UNSUPPORTED = "unsupported"
HEATING_VALUE_READY_YEAR = 2025

# High-confidence aliases are specific business labels.
# Medium-confidence aliases are usable but ambiguous and need confirmation.
COLUMN_ALIAS_RULES: dict[str, dict[str, tuple[str, ...]]] = {
    "activity_type": {
        CONFIDENCE_HIGH: (
            "activity_type",
            "activity type",
            "活動類型",
            "能源別",
            "能源種類",
            "能源項目",
            "energy",
            "energy type",
            "fuel type",
        ),
        CONFIDENCE_MEDIUM: (
            "activity",
            "type",
            "項目",
            "類別",
            "活動",
        ),
    },
    "activity_value": {
        CONFIDENCE_HIGH: (
            "activity_value",
            "activity value",
            "活動量",
            "使用量",
            "耗用量",
            "消耗量",
            "能源使用量",
            "consumption",
            "usage",
            "amount",
            "quantity",
            "value",
        ),
        CONFIDENCE_MEDIUM: (
            "數量",
            "用量",
        ),
    },
    "unit": {
        CONFIDENCE_HIGH: (
            "unit",
            "單位",
            "計量單位",
            "uom",
            "unit of measure",
        ),
        CONFIDENCE_MEDIUM: (),
    },
    "site_id": {
        CONFIDENCE_HIGH: (
            "site_id",
            "site",
            "廠區",
            "場址",
            "據點",
            "工廠",
            "廠別",
            "廠場",
            "營運據點",
            "location",
            "plant",
            "facility",
        ),
        CONFIDENCE_MEDIUM: (),
    },
    "year_month": {
        CONFIDENCE_HIGH: (
            "年月",
            "月份",
            "year month",
            "year_month",
            "year-month",
            "month",
        ),
        CONFIDENCE_MEDIUM: (
            "期間",
            "period",
            "日期",
            "date",
        ),
    },
    "activity_start_date": {
        CONFIDENCE_HIGH: (
            "activity_start_date",
            "start_date",
            "start date",
            "period_start",
            "period start",
            "開始日期",
            "起始日期",
            "期間開始",
        ),
        CONFIDENCE_MEDIUM: (),
    },
    "activity_end_date": {
        CONFIDENCE_HIGH: (
            "activity_end_date",
            "end_date",
            "end date",
            "period_end",
            "period end",
            "結束日期",
            "截止日期",
            "期間結束",
            "結束日",
        ),
        CONFIDENCE_MEDIUM: (),
    },
    "fuel_subtype": {
        CONFIDENCE_HIGH: (
            "fuel_subtype",
            "natural_gas_type",
            "ng_type",
            "ng type",
            "天然氣類型",
            "天然氣種類",
            "NG類型",
            "NG 類型",
        ),
        CONFIDENCE_MEDIUM: (),
    },
}

# Backward-compatible flat alias map used by older call sites / docs.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    field_name: rules.get(CONFIDENCE_HIGH, ()) + rules.get(CONFIDENCE_MEDIUM, ())
    for field_name, rules in COLUMN_ALIAS_RULES.items()
}

# Uploaded calculation-like columns are source/reference only.
REFERENCE_ONLY_ALIASES: tuple[str, ...] = (
    "排放係數",
    "排放量",
    "排放量 (kgco2e)",
    "排放量(kgco2e)",
    "kgco2e",
    "co2e",
    "計算結果",
    "emission factor",
    "emission_factor",
    "emissions",
    "calculated emissions",
    "calculation result",
)

ACTIVITY_VALUE_ALIASES: dict[str, str] = {
    "electricity": "grid_electricity",
    "grid electricity": "grid_electricity",
    "grid_electricity": "grid_electricity",
    "外購電力": "grid_electricity",
    "電力": "grid_electricity",
    "用電": "grid_electricity",
    "natural gas": "natural_gas",
    "natural_gas": "natural_gas",
    "天然氣": "natural_gas",
    "ng1": "natural_gas",
    "ng2": "natural_gas",
    "diesel": "diesel",
    "柴油": "diesel",
    "公司車輛柴油": "diesel",
    "公務車柴油": "diesel",
    "company vehicle diesel": "diesel",
    "steel": "purchased_steel",
    "purchased steel": "purchased_steel",
    "purchased_steel": "purchased_steel",
    "鋼材": "purchased_steel",
    "採購鋼材": "purchased_steel",
    "盤元": "purchased_steel",
    "production": "finished_goods_output",
    "output": "finished_goods_output",
    "finished_goods_output": "finished_goods_output",
    "成品": "finished_goods_output",
    "產量": "finished_goods_output",
    "生產數量": "finished_goods_output",
    "third_party_transport": "third_party_transport",
    "scrap_output": "scrap_output",
    "other": "other",
    "unknown": "unknown",
}

UNIT_VALUE_ALIASES: dict[str, str] = {
    "kwh": "kWh",
    "度": "kWh",
    "mwh": "MWh",
    "m3": "m3",
    "m³": "m3",
    "立方公尺": "m3",
    "立方米": "m3",
    "l": "L",
    "公升": "L",
    "liter": "L",
    "litre": "L",
    "kg": "kg",
    "公斤": "kg",
    "t": "t",
    "噸": "t",
    "公噸": "t",
    "tonne": "t",
    "metric ton": "t",
}

RECORD_TYPE_BY_ACTIVITY: dict[str, str] = {
    "grid_electricity": "emission_activity",
    "natural_gas": "emission_activity",
    "diesel": "emission_activity",
    "purchased_steel": "material_input",
    "third_party_transport": "transport_activity",
    "finished_goods_output": "production_output",
    "scrap_output": "scrap_output",
    "other": "other",
    "unknown": "unknown",
}

REJECTION_COLUMNS = (
    "source_row",
    "field",
    "issue_code",
    "issue_message",
    "uploaded_value",
)

HEADER_SCAN_ROWS = 15
_YEAR_MONTH_RE = re.compile(
    r"^\s*(?P<year>\d{4})\s*[-/年.]\s*(?P<month>\d{1,2})\s*月?\s*$"
)
_PROSE_HINTS = (
    "說明",
    "注意",
    "請先",
    "本表",
    "填寫說明",
    "instruction",
    "readme",
    "help",
    "note:",
    "請閱讀",
)
_ACTIVITY_TEXT_HINTS = (
    "電力",
    "天然氣",
    "柴油",
    "能源",
    "electric",
    "gas",
    "diesel",
    "fuel",
    "steel",
    "鋼",
)
_UNIT_TEXT_HINTS = (
    "kwh",
    "mwh",
    "m3",
    "m³",
    "kg",
    "公升",
    "噸",
    "度",
)
_WEAK_POSITIVE_SHEET_NAMES = {
    "活動數據": 3,
    "活動資料": 3,
    "data": 2,
    "sheet1": 1,
    "january": 1,
}
_WEAK_NEGATIVE_SHEET_NAMES = {
    "說明": -4,
    "readme": -4,
    "help": -4,
    "instruction": -4,
    "instructions": -4,
}


@dataclass(frozen=True)
class UploadedTable:
    """In-memory parsed table from an uploaded CSV/XLSX file."""

    file_name: str
    file_extension: str
    sha256: str
    sheet_name: str | None
    sheet_names: tuple[str, ...]
    columns: tuple[str, ...]
    frame: pd.DataFrame
    byte_length: int
    header_row_index: int = 0
    header_needs_confirmation: bool = False


@dataclass
class ColumnMapping:
    """User-confirmed column and value mappings for one upload."""

    activity_type_column: str = ""
    activity_value_column: str = ""
    unit_column: str = ""
    site_column: str = ""
    use_file_dates: bool = True
    use_year_month: bool = False
    year_month_column: str = ""
    year_month_confirmed: bool = False
    start_date_column: str = ""
    end_date_column: str = ""
    period_start: date | None = None
    period_end: date | None = None
    natural_gas_subtype: str = "unknown"
    natural_gas_subtype_column: str = ""
    diesel_context: str = "unknown"
    electricity_context: str = "unknown"
    activity_type_value_map: dict[str, str] = field(default_factory=dict)
    unit_value_map: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class FieldSuggestion:
    """One suggested mapping from an uploaded column to a semantic field."""

    field: str
    source_column: str
    confidence: str


@dataclass(frozen=True)
class WorksheetScore:
    """Deterministic structural score for one worksheet."""

    sheet_name: str
    score: float
    data_row_count: int
    column_count: int
    has_numeric: bool
    has_date_like: bool
    has_unit_like: bool
    has_activity_like: bool
    signal_summary: tuple[str, ...]


@dataclass(frozen=True)
class HeaderDetectionResult:
    """Result of scanning the first rows for a plausible header."""

    header_row_index: int
    confidence: str
    needs_confirmation: bool
    candidate_rows: tuple[int, ...]
    preview: tuple[tuple[str, ...], ...]


@dataclass
class IntakeMetadata:
    """Once-per-upload metadata provided by the beginner."""

    source_name: str
    site_id: str
    document_date: date | None
    data_quality_tier: str
    intake_run_id: str
    ingested_at: pd.Timestamp


@dataclass
class IntakeValidationResult:
    """Accepted / rejected preview after canonical build + schema checks."""

    source_documents: pd.DataFrame
    accepted_activities: pd.DataFrame
    rejected_rows: pd.DataFrame
    accepted_count: int
    rejected_count: int
    total_count: int
    file_hash: str
    file_name: str
    potential_duplicate_groups: tuple[Any, ...] = ()


class IntakeError(ValueError):
    """Beginner-facing intake failure with a stable issue code."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


def sanitize_filename(name: str) -> str:
    """Return basename only, stripping path components and control chars."""
    raw = str(name or "").strip()
    base = Path(raw.replace("\\", "/")).name
    cleaned = re.sub(r"[\x00-\x1f\x7f]", "", base).strip()
    return cleaned or "uploaded_file"


def compute_bytes_sha256(data: bytes) -> str:
    """Return lowercase SHA-256 hex digest of exact bytes."""
    return hashlib.sha256(data).hexdigest()


def validate_upload_bytes(file_name: str, data: bytes) -> str:
    """Validate extension and size; return sanitized lowercase extension."""
    safe_name = sanitize_filename(file_name)
    extension = Path(safe_name).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise IntakeError(
            ISSUE_UNSUPPORTED_FILE_TYPE,
            f"Unsupported file type: {extension or '(none)'}.",
        )
    if len(data) > MAX_UPLOAD_BYTES:
        raise IntakeError(
            ISSUE_FILE_TOO_LARGE,
            "File exceeds the 10 MB Phase 9A limit.",
        )
    if not data:
        raise IntakeError(
            ISSUE_INVALID_ENCODING,
            "Uploaded file is empty.",
        )
    return extension


def blank_template_csv_bytes() -> bytes:
    """Return UTF-8 CSV bytes for the blank downloadable template."""
    header = ",".join(BLANK_TEMPLATE_COLUMNS) + "\n"
    return header.encode("utf-8")


def example_csv_bytes() -> bytes:
    """Return UTF-8 CSV bytes for the beginner example download."""
    lines = [
        ",".join(BLANK_TEMPLATE_COLUMNS),
        "外購電力,50000,kWh,2024-01-01,2024-01-31",
        "天然氣,8000,m3,2024-01-01,2024-01-31",
        "柴油,1200,L,2024-01-01,2024-01-31",
        "採購鋼材,150,t,2024-01-01,2024-01-31",
        "",
    ]
    return "\n".join(lines).encode("utf-8")


def example_preview_rows() -> pd.DataFrame:
    """On-screen example rows that are never imported."""
    return pd.DataFrame(
        [
            {
                "activity_type": "外購電力",
                "activity_value": 50000,
                "unit": "kWh",
                "activity_start_date": "2024-01-01",
                "activity_end_date": "2024-01-31",
            },
            {
                "activity_type": "天然氣",
                "activity_value": 8000,
                "unit": "m3",
                "activity_start_date": "2024-01-01",
                "activity_end_date": "2024-01-31",
            },
            {
                "activity_type": "柴油",
                "activity_value": 1200,
                "unit": "L",
                "activity_start_date": "2024-01-01",
                "activity_end_date": "2024-01-31",
            },
            {
                "activity_type": "採購鋼材",
                "activity_value": 150,
                "unit": "t",
                "activity_start_date": "2024-01-01",
                "activity_end_date": "2024-01-31",
            },
        ]
    )


CUSTOMER_TEMPLATE_HEADERS = (
    "活動類型",
    "用量",
    "單位",
    "開始日期",
    "結束日期",
)


def example_preview_customer_rows() -> pd.DataFrame:
    """Example rows with customer-facing column names (never imported)."""
    frame = example_preview_rows()
    return frame.rename(
        columns={
            "activity_type": CUSTOMER_TEMPLATE_HEADERS[0],
            "activity_value": CUSTOMER_TEMPLATE_HEADERS[1],
            "unit": CUSTOMER_TEMPLATE_HEADERS[2],
            "activity_start_date": CUSTOMER_TEMPLATE_HEADERS[3],
            "activity_end_date": CUSTOMER_TEMPLATE_HEADERS[4],
        }
    )


_BLANK_TEMPLATE_XLSX_CACHE: bytes | None = None


def blank_template_xlsx_bytes() -> bytes:
    """Workbook: fill sheet, example sheet, field guide. Never auto-imported."""
    global _BLANK_TEMPLATE_XLSX_CACHE
    if _BLANK_TEMPLATE_XLSX_CACHE is None:
        _BLANK_TEMPLATE_XLSX_CACHE = _build_blank_template_xlsx_bytes()
    return _BLANK_TEMPLATE_XLSX_CACHE


def _build_blank_template_xlsx_bytes() -> bytes:
    from datetime import datetime, timezone
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    # Stable timestamps so Streamlit download_button data hashes stay constant.
    workbook.properties.created = datetime(2024, 1, 1, tzinfo=timezone.utc)
    workbook.properties.modified = datetime(2024, 1, 1, tzinfo=timezone.utc)

    fill_sheet = workbook.active
    fill_sheet.title = "資料填寫"
    fill_sheet.append(list(CUSTOMER_TEMPLATE_HEADERS))

    example_sheet = workbook.create_sheet("填寫範例")
    example_sheet.append(list(CUSTOMER_TEMPLATE_HEADERS))
    for row in example_preview_customer_rows().itertuples(index=False):
        example_sheet.append(
            [
                str(row[0]),
                int(row[1]),
                str(row[2]),
                str(row[3]),
                str(row[4]),
            ]
        )

    guide = workbook.create_sheet("欄位說明")
    guide.append(["欄位", "說明", "English", "系統欄位名稱"])
    guide.append(
        ["活動類型", "這筆資料是哪一種活動", "Activity type", "activity_type"]
    )
    guide.append(["用量", "實際使用或採購數量", "Quantity", "activity_value"])
    guide.append(["單位", "例如 kWh、m3、L、t", "Unit", "unit"])
    guide.append(
        ["開始日期", "資料期間的第一天", "Start date", "activity_start_date"]
    )
    guide.append(
        ["結束日期", "資料期間的最後一天", "End date", "activity_end_date"]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_HEADER_PAREN_RE = re.compile(r"[\(\[（【].*?[\)\]）】]")
_HEADER_WRAPPER_SUFFIXES = ("欄位", "欄", "column", "col")
_VALUE_SAMPLE_LIMIT = 30


def _normalize_header(value: str) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def structural_normalize_header(value: str) -> str:
    """NFKC header identity: strip parenthetical units and collapse space."""
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = _HEADER_PAREN_RE.sub(" ", text)
    text = text.replace("_", " ").replace("-", " ").replace(":", " ")
    text = text.strip().lower()
    text = re.sub(r"[^\w\u4e00-\u9fff ]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def is_reference_only_column(column_name: str) -> bool:
    """Return True when a column looks like uploaded calc/factor output."""
    normalized = _normalize_header(column_name)
    if not normalized:
        return False
    for alias in REFERENCE_ONLY_ALIASES:
        alias_norm = _normalize_header(alias)
        if normalized == alias_norm or alias_norm in normalized:
            return True
    if "排放" in str(column_name) and (
        "係數" in str(column_name) or "量" in str(column_name)
    ):
        return True
    return False


def reference_only_columns(columns: list[str] | tuple[str, ...]) -> list[str]:
    """Return uploaded columns treated as source/reference only."""
    return [col for col in columns if is_reference_only_column(col)]


def _alias_match_confidence(column_name: str, field_name: str) -> str:
    if is_reference_only_column(column_name):
        return CONFIDENCE_LOW
    normalized = structural_normalize_header(column_name)
    if not normalized:
        return CONFIDENCE_LOW
    rules = COLUMN_ALIAS_RULES.get(field_name, {})
    for alias in rules.get(CONFIDENCE_HIGH, ()):
        if normalized == structural_normalize_header(alias):
            return CONFIDENCE_HIGH
    for alias in rules.get(CONFIDENCE_MEDIUM, ()):
        if normalized == structural_normalize_header(alias):
            return CONFIDENCE_MEDIUM
    return CONFIDENCE_LOW


def _wrapped_high_alias_match(column_name: str, field_name: str) -> bool:
    """Header equals a High alias plus a short wrapper. Always Medium, never High."""
    if is_reference_only_column(column_name):
        return False
    normalized = structural_normalize_header(column_name)
    if not normalized:
        return False
    rules = COLUMN_ALIAS_RULES.get(field_name, {})
    for alias in rules.get(CONFIDENCE_HIGH, ()):
        alias_norm = structural_normalize_header(alias)
        if not alias_norm or alias_norm == normalized:
            continue
        if normalized.startswith(alias_norm + " "):
            suffix = normalized[len(alias_norm) :].strip()
            if suffix in _HEADER_WRAPPER_SUFFIXES:
                return True
        if normalized.endswith(" " + alias_norm):
            prefix = normalized[: -len(alias_norm)].strip()
            if prefix in {"本月", "本期"}:
                return True
        compact = normalized.replace(" ", "")
        alias_compact = alias_norm.replace(" ", "")
        for suffix in _HEADER_WRAPPER_SUFFIXES:
            if compact == alias_compact + suffix.replace(" ", ""):
                return True
    return False


def _sampled_texts(frame: pd.DataFrame, column: str) -> list[str]:
    if column not in frame.columns:
        return []
    texts: list[str] = []
    for value in frame[column].tolist():
        text = _cell_text(value)
        if not text:
            continue
        texts.append(text)
        if len(texts) >= _VALUE_SAMPLE_LIMIT:
            break
    return texts


def _value_assisted_field(
    field_name: str,
    texts: list[str],
    *,
    facility_names: tuple[str, ...] = (),
) -> bool:
    if not texts:
        return False
    total = len(texts)
    if field_name == "activity_value":
        numeric = sum(1 for item in texts if _is_numeric_cell(item))
        dated = sum(
            1
            for item in texts
            if _is_date_like_cell(item) and not _is_numeric_cell(item)
        )
        return numeric / total >= 0.7 and dated / total < 0.3
    if field_name == "unit":
        units = sum(1 for item in texts if bool(suggest_unit(item)))
        return units / total >= 0.6
    if field_name == "activity_type":
        matched = sum(1 for item in texts if bool(suggest_activity_type(item)))
        return matched / total >= 0.6
    if field_name in {"activity_start_date", "activity_end_date", "year_month"}:
        dated = sum(1 for item in texts if _is_date_like_cell(item))
        return dated / total >= 0.6
    if field_name == "site_id" and facility_names:
        wanted = {structural_normalize_header(name) for name in facility_names}
        wanted.discard("")
        if not wanted:
            return False
        hits = sum(
            1 for item in texts if structural_normalize_header(item) in wanted
        )
        return hits / total >= 0.5
    return False


def suggest_column_mapping_with_confidence(
    columns: list[str] | tuple[str, ...],
    *,
    frame: pd.DataFrame | None = None,
    facility_names: tuple[str, ...] | list[str] = (),
) -> dict[str, FieldSuggestion]:
    """Suggest semantic mappings with confidence; never claim reference cols.

    Exact unambiguous alias matches may be High. Fuzzy, value-assisted, or
    competing matches remain Medium or Low and are never silently committed.
    """
    usable = [col for col in columns if not is_reference_only_column(col)]
    claimed: set[str] = set()
    suggestions: dict[str, FieldSuggestion] = {}
    facilities = tuple(facility_names or ())

    # Prefer start/end date fields over ambiguous year-month when both exist.
    field_order = (
        "activity_type",
        "activity_value",
        "unit",
        "site_id",
        "activity_start_date",
        "activity_end_date",
        "year_month",
    )

    for field_name in field_order:
        high_cols: list[str] = []
        medium_cols: list[str] = []
        for col in usable:
            if col in claimed:
                continue
            confidence = _alias_match_confidence(col, field_name)
            if confidence == CONFIDENCE_HIGH:
                high_cols.append(col)
            elif confidence == CONFIDENCE_MEDIUM:
                medium_cols.append(col)
            elif _wrapped_high_alias_match(col, field_name):
                medium_cols.append(col)
        best_col = ""
        best_confidence = CONFIDENCE_LOW
        if len(high_cols) == 1:
            best_col = high_cols[0]
            best_confidence = CONFIDENCE_HIGH
        elif len(high_cols) > 1:
            best_col = high_cols[0]
            best_confidence = CONFIDENCE_MEDIUM
        elif len(medium_cols) == 1:
            best_col = medium_cols[0]
            best_confidence = CONFIDENCE_MEDIUM
        elif len(medium_cols) > 1:
            best_col = medium_cols[0]
            best_confidence = CONFIDENCE_MEDIUM
        elif frame is not None:
            assisted = [
                col
                for col in usable
                if col not in claimed
                and _value_assisted_field(
                    field_name,
                    _sampled_texts(frame, col),
                    facility_names=facilities,
                )
            ]
            if len(assisted) == 1:
                best_col = assisted[0]
                best_confidence = CONFIDENCE_MEDIUM
        if best_col and best_confidence != CONFIDENCE_LOW:
            suggestions[field_name] = FieldSuggestion(
                field=field_name,
                source_column=best_col,
                confidence=best_confidence,
            )
            claimed.add(best_col)
        else:
            suggestions[field_name] = FieldSuggestion(
                field=field_name,
                source_column="",
                confidence=CONFIDENCE_LOW,
            )

    # If both start and end are present, drop year_month auto-suggestion.
    start_suggestion = suggestions.get(
        "activity_start_date",
        FieldSuggestion("", "", ""),
    )
    end_suggestion = suggestions.get(
        "activity_end_date",
        FieldSuggestion("", "", ""),
    )
    if start_suggestion.source_column and end_suggestion.source_column:
        ym = suggestions.get("year_month")
        if ym and ym.source_column:
            suggestions["year_month"] = FieldSuggestion(
                field="year_month",
                source_column="",
                confidence=CONFIDENCE_LOW,
            )

    return suggestions


def suggest_column_mapping(columns: list[str] | tuple[str, ...]) -> dict[str, str]:
    """Deterministic column suggestions from aliases; empty when unmatched.

    High- and medium-confidence matches are returned for user confirmation.
    Low-confidence matches remain empty (never auto-selected).
    """
    detailed = suggest_column_mapping_with_confidence(columns)
    return {
        field_name: (
            suggestion.source_column
            if suggestion.confidence in {CONFIDENCE_HIGH, CONFIDENCE_MEDIUM}
            else ""
        )
        for field_name, suggestion in detailed.items()
    }


def extract_natural_gas_subtype_from_text(value: Any) -> str:
    """Return NG1/NG2 when the source cell explicitly names the type."""
    text = str(value if value is not None else "").strip()
    if not text:
        return ""
    match = _NG_SUBTYPE_RE.search(text.replace("　", " "))
    if match:
        return f"NG{match.group(1)}"
    compact = re.sub(r"\s+", "", text).upper()
    if compact in {"NG1", "NG2"}:
        return compact
    return ""


def extract_diesel_vehicle_context_from_text(value: Any) -> bool:
    """True when source text explicitly indicates company-vehicle diesel."""
    text = str(value if value is not None else "").strip().lower()
    if not text:
        return False
    return any(hint in text for hint in _DIESEL_VEHICLE_HINTS)


def suggest_activity_type(value: Any) -> str:
    """Suggest a canonical activity type; empty string when unmatched."""
    text = str(value if value is not None else "").strip()
    if not text:
        return UNMAPPED_SENTINEL
    direct = ACTIVITY_VALUE_ALIASES.get(text.lower())
    if direct:
        return direct
    zh = ACTIVITY_VALUE_ALIASES.get(text)
    if zh:
        return zh
    if extract_natural_gas_subtype_from_text(text) or "天然氣" in text:
        return "natural_gas"
    lowered = text.lower()
    if "natural gas" in lowered:
        return "natural_gas"
    if "電力" in text or "electricity" in lowered:
        return "grid_electricity"
    if "柴油" in text or "diesel" in lowered:
        return "diesel"
    if "鋼" in text or "steel" in lowered:
        return "purchased_steel"
    return UNMAPPED_SENTINEL


def suggest_unit(value: Any) -> str:
    """Suggest a canonical unit; empty string when unmatched."""
    text = str(value if value is not None else "").strip()
    if not text:
        return UNMAPPED_SENTINEL
    if text in SUPPORTED_UNITS:
        return text
    return UNIT_VALUE_ALIASES.get(text.lower(), UNMAPPED_SENTINEL)


def distinct_values(frame: pd.DataFrame, column: str) -> list[str]:
    """Return sorted distinct non-null string values for a column."""
    if not column or column not in frame.columns:
        return []
    series = frame[column]
    values: list[str] = []
    seen: set[str] = set()
    for item in series.tolist():
        if item is None or (isinstance(item, float) and math.isnan(item)):
            continue
        text = str(item).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        values.append(text)
    return sorted(values)


def parse_year_month_period(value: Any) -> tuple[date, date]:
    """Parse a year-month label into inclusive calendar month bounds.

    Supports forms such as ``2025-01``, ``2025/01``, and ``2025年1月``.
    Leap years are handled via ``calendar.monthrange``.
    """
    if value is None or (isinstance(value, float) and math.isnan(value)):
        raise ValueError("year-month value is missing")
    if isinstance(value, datetime):
        year, month = value.year, value.month
    elif isinstance(value, date):
        year, month = value.year, value.month
    else:
        text = str(value).strip()
        if not text:
            raise ValueError("year-month value is missing")
        # Excel serial / Timestamp-like via pandas when unambiguous full date.
        match = _YEAR_MONTH_RE.match(text)
        if match:
            year = int(match.group("year"))
            month = int(match.group("month"))
        else:
            parsed = pd.to_datetime(text, errors="coerce")
            if pd.isna(parsed):
                raise ValueError(f"unrecognized year-month value: {text!r}")
            timestamp = pd.Timestamp(parsed)
            year, month = int(timestamp.year), int(timestamp.month)
    if month < 1 or month > 12:
        raise ValueError(f"invalid month in year-month value: {month}")
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def year_month_transform_preview(value: Any) -> dict[str, str]:
    """Return a beginner-facing preview of a confirmed year-month transform."""
    start, end = parse_year_month_period(value)
    return {
        "source": str(value).strip(),
        "activity_start_date": start.isoformat(),
        "activity_end_date": end.isoformat(),
    }


def list_xlsx_sheet_names(data: bytes) -> list[str]:
    """Return workbook sheet names without writing to disk."""
    frame_dict = pd.read_excel(
        BytesIO(data),
        sheet_name=None,
        header=None,
        dtype=object,
        engine="openpyxl",
    )
    return list(frame_dict.keys())


def _cell_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip()


def _is_numeric_cell(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return math.isfinite(float(value))
    text = _cell_text(value).replace(",", "")
    if not text:
        return False
    try:
        number = float(text)
    except ValueError:
        return False
    return math.isfinite(number)


def _is_date_like_cell(value: Any) -> bool:
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return True
    text = _cell_text(value)
    if not text:
        return False
    if _YEAR_MONTH_RE.match(text):
        return True
    parsed = pd.to_datetime(text, errors="coerce")
    return not pd.isna(parsed)


def _row_nonempty_count(row: pd.Series) -> int:
    return sum(1 for value in row.tolist() if _cell_text(value))


def _row_looks_like_header(row: pd.Series) -> bool:
    values = [_cell_text(value) for value in row.tolist()]
    nonempty = [value for value in values if value]
    if len(nonempty) < 2:
        return False
    textish = 0
    numericish = 0
    long_prose = 0
    for value in nonempty:
        if _is_numeric_cell(value):
            numericish += 1
        else:
            textish += 1
        if len(value) > 40:
            long_prose += 1
    if long_prose >= max(1, len(nonempty) // 2):
        return False
    if textish < max(2, len(nonempty) - 1):
        return False
    if numericish > textish:
        return False
    return True


def detect_header_row(raw_frame: pd.DataFrame) -> HeaderDetectionResult:
    """Detect the most plausible header row in the first scan window."""
    if raw_frame is None or raw_frame.empty:
        return HeaderDetectionResult(
            header_row_index=0,
            confidence=CONFIDENCE_LOW,
            needs_confirmation=True,
            candidate_rows=(0,),
            preview=tuple(),
        )

    limit = min(HEADER_SCAN_ROWS, len(raw_frame))
    scored: list[tuple[float, int]] = []
    for index in range(limit):
        row = raw_frame.iloc[index]
        nonempty = _row_nonempty_count(row)
        if nonempty < 2:
            continue
        score = float(nonempty * 3)
        if _row_looks_like_header(row):
            score += 25
        # Prefer rows followed by tabular data.
        following = raw_frame.iloc[index + 1 : min(index + 4, len(raw_frame))]
        if not following.empty:
            numeric_follow = 0
            populated_follow = 0
            for _, follow_row in following.iterrows():
                vals = [_cell_text(v) for v in follow_row.tolist()]
                populated = [v for v in vals if v]
                if len(populated) >= 2:
                    populated_follow += 1
                numeric_follow += sum(1 for v in populated if _is_numeric_cell(v))
            score += populated_follow * 6
            score += min(numeric_follow, 6) * 2
        # Penalize report-title / prose rows.
        joined = " ".join(_cell_text(v) for v in row.tolist())
        if any(hint in joined for hint in _PROSE_HINTS):
            score -= 20
        if nonempty <= 2 and any(len(_cell_text(v)) > 30 for v in row.tolist()):
            score -= 15
        # Alias hits are a strong positive signal.
        alias_hits = 0
        for value in row.tolist():
            text = _cell_text(value)
            if not text or is_reference_only_column(text):
                continue
            for field_name in COLUMN_ALIAS_RULES:
                if _alias_match_confidence(text, field_name) != CONFIDENCE_LOW:
                    alias_hits += 1
                    break
        score += alias_hits * 8
        scored.append((score, index))

    if not scored:
        preview = tuple(
            tuple(_cell_text(v) for v in raw_frame.iloc[i].tolist())
            for i in range(min(5, len(raw_frame)))
        )
        return HeaderDetectionResult(
            header_row_index=0,
            confidence=CONFIDENCE_LOW,
            needs_confirmation=True,
            candidate_rows=(0,),
            preview=preview,
        )

    scored.sort(key=lambda item: (-item[0], item[1]))
    best_score, best_index = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else float("-inf")
    needs_confirmation = False
    confidence = CONFIDENCE_HIGH
    if best_score < 20:
        confidence = CONFIDENCE_LOW
        needs_confirmation = True
    elif best_score - second_score < 8:
        confidence = CONFIDENCE_MEDIUM
        needs_confirmation = True
    elif best_index != 0 and best_score < 35:
        confidence = CONFIDENCE_MEDIUM
        needs_confirmation = True

    candidate_rows = tuple(index for _, index in scored[:5])
    preview_indices = sorted(set(candidate_rows[:3]) | {best_index})
    preview = tuple(
        tuple(_cell_text(v) for v in raw_frame.iloc[i].tolist())
        for i in preview_indices
        if i < len(raw_frame)
    )
    return HeaderDetectionResult(
        header_row_index=best_index,
        confidence=confidence,
        needs_confirmation=needs_confirmation,
        candidate_rows=candidate_rows,
        preview=preview,
    )


def _frame_signal_flags(raw_frame: pd.DataFrame, header_index: int) -> dict[str, Any]:
    if raw_frame.empty:
        return {
            "data_row_count": 0,
            "column_count": 0,
            "has_numeric": False,
            "has_date_like": False,
            "has_unit_like": False,
            "has_activity_like": False,
            "prose_heavy": True,
            "alias_hits": 0,
        }
    header = raw_frame.iloc[header_index]
    body = raw_frame.iloc[header_index + 1 :]
    headers = [_cell_text(v) for v in header.tolist()]
    usable_cols = [h for h in headers if h]
    data_row_count = 0
    numeric_cells = 0
    date_like = 0
    unit_like = 0
    activity_like = 0
    long_text = 0
    populated_cells = 0
    for _, row in body.iterrows():
        values = [_cell_text(v) for v in row.tolist()]
        nonempty = [v for v in values if v]
        if len(nonempty) < 2:
            continue
        data_row_count += 1
        for value in nonempty:
            populated_cells += 1
            if _is_numeric_cell(value):
                numeric_cells += 1
            if _is_date_like_cell(value):
                date_like += 1
            lowered = value.lower()
            if any(hint in lowered for hint in _UNIT_TEXT_HINTS) or value in {
                "L",
                "t",
                "kg",
                "kWh",
                "MWh",
                "m3",
            }:
                unit_like += 1
            if any(hint in value for hint in _ACTIVITY_TEXT_HINTS) or any(
                hint in lowered for hint in _ACTIVITY_TEXT_HINTS
            ):
                activity_like += 1
            if len(value) > 48:
                long_text += 1

    alias_hits = 0
    for header_text in usable_cols:
        if is_reference_only_column(header_text):
            continue
        for field_name in COLUMN_ALIAS_RULES:
            if _alias_match_confidence(header_text, field_name) != CONFIDENCE_LOW:
                alias_hits += 1
                break

    prose_heavy = (
        data_row_count <= 1
        and long_text >= max(1, populated_cells // 2)
        and numeric_cells == 0
    ) or (
        populated_cells > 0
        and long_text >= max(2, populated_cells // 2)
        and numeric_cells == 0
        and data_row_count <= 3
    )
    return {
        "data_row_count": data_row_count,
        "column_count": len(usable_cols),
        "has_numeric": numeric_cells > 0,
        "has_date_like": date_like > 0
        or any(
            _alias_match_confidence(h, "year_month") != CONFIDENCE_LOW
            or _alias_match_confidence(h, "activity_start_date") != CONFIDENCE_LOW
            for h in usable_cols
        ),
        "has_unit_like": unit_like > 0
        or any(
            _alias_match_confidence(h, "unit") != CONFIDENCE_LOW
            for h in usable_cols
        ),
        "has_activity_like": activity_like > 0
        or any(
            _alias_match_confidence(h, "activity_type") != CONFIDENCE_LOW
            for h in usable_cols
        ),
        "prose_heavy": prose_heavy,
        "alias_hits": alias_hits,
        "numeric_cells": numeric_cells,
        "long_text": long_text,
    }


def score_worksheet(raw_frame: pd.DataFrame, sheet_name: str) -> WorksheetScore:
    """Score one worksheet using deterministic structural signals."""
    header = detect_header_row(raw_frame)
    signals = _frame_signal_flags(raw_frame, header.header_row_index)
    score = 0.0
    summary: list[str] = []

    data_rows = int(signals["data_row_count"])
    column_count = int(signals["column_count"])
    score += min(data_rows, 40) * 2.0
    if data_rows >= 3:
        score += 12
        summary.append("multiple_data_rows")
    if column_count >= 3:
        score += 10
        summary.append("stable_columns")
    if signals["has_numeric"]:
        score += 18
        summary.append("numeric")
    if signals["has_date_like"]:
        score += 14
        summary.append("date_or_month")
    if signals["has_activity_like"]:
        score += 14
        summary.append("activity_or_energy")
    if signals["has_unit_like"]:
        score += 12
        summary.append("unit")
    score += min(int(signals["alias_hits"]), 6) * 5
    if int(signals["alias_hits"]) > 0:
        summary.append("header_aliases")

    if signals["prose_heavy"]:
        score -= 45
        summary.append("prose_heavy")
    if data_rows < 2:
        score -= 25
        summary.append("few_rows")
    if not signals["has_numeric"]:
        score -= 18
        summary.append("no_numeric")
    joined_sample = " ".join(
        _cell_text(v)
        for v in raw_frame.head(min(8, len(raw_frame))).to_numpy().ravel().tolist()
    )
    if any(hint in joined_sample for hint in _PROSE_HINTS):
        score -= 16
        summary.append("instruction_like")

    # Sheet name is only a weak supporting hint.
    name_key = _normalize_header(sheet_name)
    score += float(_WEAK_POSITIVE_SHEET_NAMES.get(name_key, 0))
    score += float(_WEAK_NEGATIVE_SHEET_NAMES.get(name_key, 0))

    return WorksheetScore(
        sheet_name=sheet_name,
        score=score,
        data_row_count=data_rows,
        column_count=column_count,
        has_numeric=bool(signals["has_numeric"]),
        has_date_like=bool(signals["has_date_like"]),
        has_unit_like=bool(signals["has_unit_like"]),
        has_activity_like=bool(signals["has_activity_like"]),
        signal_summary=tuple(summary),
    )


def rank_xlsx_worksheets(data: bytes) -> list[WorksheetScore]:
    """Rank every worksheet by structural fitness for activity intake."""
    workbook = pd.read_excel(
        BytesIO(data),
        sheet_name=None,
        header=None,
        dtype=object,
        engine="openpyxl",
    )
    ranked = [
        score_worksheet(
            frame if isinstance(frame, pd.DataFrame) else pd.DataFrame(),
            name,
        )
        for name, frame in workbook.items()
    ]
    ranked.sort(key=lambda item: (-item.score, item.sheet_name))
    return ranked


def suggest_xlsx_worksheet(data: bytes) -> WorksheetScore | None:
    """Return the top-ranked worksheet suggestion, if any."""
    ranked = rank_xlsx_worksheets(data)
    return ranked[0] if ranked else None


def worksheet_detection_labels(score: WorksheetScore) -> list[str]:
    """Beginner-facing detection bullets for a worksheet suggestion."""
    labels: list[str] = [
        f"{score.data_row_count} 筆資料",
        f"{score.column_count} 個欄位",
    ]
    bits: list[str] = []
    if score.has_numeric:
        bits.append("數量")
    if score.has_unit_like:
        bits.append("單位")
    if score.has_date_like:
        bits.append("年月")
    if score.has_activity_like:
        bits.append("能源/活動")
    if bits:
        labels.append("包含" + "、".join(bits) + "欄位")
    return labels


def _unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        text = _cell_text(value) or f"column_{index + 1}"
        count = seen.get(text, 0)
        seen[text] = count + 1
        headers.append(text if count == 0 else f"{text}_{count + 1}")
    return headers


def _frame_from_raw(
    raw_frame: pd.DataFrame,
    *,
    header_row: int | None,
) -> tuple[pd.DataFrame, int, bool]:
    detection = detect_header_row(raw_frame)
    selected = detection.header_row_index if header_row is None else int(header_row)
    if selected < 0 or (len(raw_frame) and selected >= len(raw_frame)):
        selected = 0
    if raw_frame.empty:
        return raw_frame.copy(), 0, True
    headers = _unique_headers(list(raw_frame.iloc[selected].tolist()))
    body = raw_frame.iloc[selected + 1 :].copy()
    body.columns = headers
    body = body.reset_index(drop=True)
    # Drop fully empty rows.
    if not body.empty:
        mask = body.apply(
            lambda row: any(_cell_text(v) for v in row.tolist()),
            axis=1,
        )
        body = body.loc[mask].reset_index(drop=True)
    needs = detection.needs_confirmation if header_row is None else False
    return body, selected, needs


def _read_csv_raw(data: bytes) -> pd.DataFrame:
    try:
        return pd.read_csv(
            BytesIO(data),
            header=None,
            dtype=object,
            encoding="utf-8-sig",
        )
    except UnicodeDecodeError:
        try:
            return pd.read_csv(
                BytesIO(data),
                header=None,
                dtype=object,
                encoding="utf-8",
            )
        except UnicodeDecodeError as exc:
            raise IntakeError(
                ISSUE_INVALID_ENCODING,
                "CSV must be UTF-8 encoded.",
            ) from exc


def load_raw_tabular_frame(
    *,
    data: bytes,
    file_extension: str,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Load CSV/XLSX bytes as a headerless frame for inspection."""
    extension = file_extension.lower()
    if not extension.startswith("."):
        extension = f".{extension}"
    if extension == ".csv":
        return _read_csv_raw(data)
    workbook = pd.read_excel(
        BytesIO(data),
        sheet_name=None,
        header=None,
        dtype=object,
        engine="openpyxl",
    )
    names = list(workbook.keys())
    if not names:
        return pd.DataFrame()
    selected = sheet_name if sheet_name in names else names[0]
    frame = workbook[selected]
    return frame if isinstance(frame, pd.DataFrame) else pd.DataFrame()


def parse_uploaded_table(
    *,
    file_name: str,
    data: bytes,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> UploadedTable:
    """Parse CSV/XLSX bytes into an in-memory UploadedTable."""
    extension = validate_upload_bytes(file_name, data)
    safe_name = sanitize_filename(file_name)
    digest = compute_bytes_sha256(data)
    sheet_names: tuple[str, ...] = ()
    selected_sheet: str | None = None

    if extension == ".csv":
        raw = _read_csv_raw(data)
        frame, header_index, header_needs = _frame_from_raw(raw, header_row=header_row)
    else:
        workbook = pd.read_excel(
            BytesIO(data),
            sheet_name=None,
            header=None,
            dtype=object,
            engine="openpyxl",
        )
        sheet_names = tuple(workbook.keys())
        if not sheet_names:
            raise IntakeError(
                ISSUE_UNSUPPORTED_FILE_TYPE,
                "Workbook contains no sheets.",
            )
        if sheet_name in sheet_names:
            selected_sheet = sheet_name
        else:
            ranked = [
                score_worksheet(
                    workbook[name]
                    if isinstance(workbook[name], pd.DataFrame)
                    else pd.DataFrame(),
                    name,
                )
                for name in sheet_names
            ]
            ranked.sort(key=lambda item: (-item.score, item.sheet_name))
            selected_sheet = ranked[0].sheet_name if ranked else sheet_names[0]
        raw = workbook[selected_sheet]
        if not isinstance(raw, pd.DataFrame):
            raw = pd.DataFrame()
        frame, header_index, header_needs = _frame_from_raw(raw, header_row=header_row)

    frame = frame.copy()
    frame.columns = [str(col) for col in frame.columns]
    return UploadedTable(
        file_name=safe_name,
        file_extension=extension,
        sha256=digest,
        sheet_name=selected_sheet,
        sheet_names=sheet_names,
        columns=tuple(str(col) for col in frame.columns),
        frame=frame,
        byte_length=len(data),
        header_row_index=header_index,
        header_needs_confirmation=header_needs,
    )


def source_document_id_from_hash(sha256: str) -> str:
    """Deterministic source_document_id from file hash."""
    return f"upload_{sha256[:12]}"


def activity_record_id(sha256: str, source_row: int) -> str:
    """Deterministic activity record_id from hash + source row number."""
    return f"up_{sha256[:12]}_r{source_row:04d}"


def source_locator(*, sheet_name: str | None, source_row: int) -> str:
    """Deterministic provenance locator back to the uploaded table."""
    if sheet_name:
        return f"sheet:{sheet_name},row:{source_row}"
    return f"row:{source_row}"


def _parse_date(value: Any, *, field_name: str) -> datetime:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        raise ValueError(f"{field_name} is missing")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    if not text:
        raise ValueError(f"{field_name} is missing")
    parsed = pd.to_datetime(text, errors="raise")
    if isinstance(parsed, pd.Timestamp):
        return parsed.to_pydatetime().replace(tzinfo=None)
    raise ValueError(f"{field_name} is invalid")


def _parse_activity_value(value: Any) -> float:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        raise ValueError("activity value is missing")
    text = str(value).strip().replace(",", "")
    if not text:
        raise ValueError("activity value is missing")
    number = float(text)
    if not math.isfinite(number):
        raise ValueError("activity value must be finite")
    if number <= 0:
        raise ValueError("activity value must be greater than zero")
    return number


def record_type_for_activity(activity_type: str) -> str:
    """Derive record_type from mapped activity_type."""
    return RECORD_TYPE_BY_ACTIVITY.get(activity_type, "unknown")


def _ownership_for_activity(activity_type: str) -> str:
    if activity_type in {
        "grid_electricity",
        "natural_gas",
        "diesel",
        "third_party_transport",
    }:
        return "unknown"
    if activity_type in {
        "purchased_steel",
        "finished_goods_output",
        "scrap_output",
    }:
        return "not_applicable"
    return "unknown"


def _activity_year_from_dates(start_dt: Any, end_dt: Any) -> int | None:
    for value in (start_dt, end_dt):
        if value is None:
            continue
        year = getattr(value, "year", None)
        if year is not None:
            try:
                return int(year)
            except (TypeError, ValueError):
                continue
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.notna(parsed):
            return int(parsed.year)
    return None


def _resolve_natural_gas_subtype(
    *,
    row: pd.Series,
    mapping: ColumnMapping,
    activity_source: Any,
) -> str:
    column = str(mapping.natural_gas_subtype_column or "").strip()
    if column and column in row.index:
        extracted = extract_natural_gas_subtype_from_text(row.get(column))
        if extracted:
            return extracted
    extracted = extract_natural_gas_subtype_from_text(activity_source)
    if extracted:
        return extracted
    ng_type = str(mapping.natural_gas_subtype or "").strip()
    if ng_type in {"NG1", "NG2"}:
        return ng_type
    return "unknown"


def _resolve_process_use(
    *,
    mapped_activity: str,
    mapping: ColumnMapping,
    activity_source: Any,
) -> str:
    if mapped_activity == "natural_gas":
        # V1 company NG activity uses the existing stationary-combustion path.
        # NG1/NG2 is never inferred here.
        return "heat_treatment"
    if mapped_activity == "diesel":
        if extract_diesel_vehicle_context_from_text(activity_source):
            return "company_vehicle"
        if str(mapping.diesel_context or "").strip() == "company_vehicle":
            return "company_vehicle"
        return "unknown"
    if mapped_activity == "grid_electricity":
        if str(mapping.electricity_context or "").strip() == "enterprise":
            return "general_factory"
        return "unknown"
    if mapped_activity in {"purchased_steel", "finished_goods_output", "scrap_output"}:
        return "not_applicable"
    return "unknown"


def _inventory_ownership(mapped_activity: str, process_use: str) -> str:
    if mapped_activity == "purchased_steel":
        return "not_applicable"
    if mapped_activity == "natural_gas":
        return "owned"
    if mapped_activity == "diesel" and process_use == "company_vehicle":
        return "owned"
    if (
        mapped_activity == "grid_electricity"
        and process_use in _ENTERPRISE_ELECTRICITY_USES
    ):
        return "owned"
    return _ownership_for_activity(mapped_activity)


def _inventory_org_boundary(mapped_activity: str, process_use: str) -> str:
    if mapped_activity == "natural_gas":
        return "inside"
    if mapped_activity == "diesel" and process_use == "company_vehicle":
        return "inside"
    if (
        mapped_activity == "grid_electricity"
        and process_use in _ENTERPRISE_ELECTRICITY_USES
    ):
        return "inside"
    return "unknown"


def classify_activity_analysis_readiness(
    *,
    activity_type: str,
    fuel_subtype: str,
    process_use: str,
    activity_start: Any,
    activity_end: Any,
) -> str:
    """Classify one accepted row for the pre-analysis business summary."""
    year = _activity_year_from_dates(activity_start, activity_end)
    if activity_type == "purchased_steel":
        return READINESS_UNSUPPORTED
    if activity_type == "natural_gas":
        if year is not None and year < HEATING_VALUE_READY_YEAR:
            return READINESS_UNSUPPORTED
        if str(fuel_subtype or "").strip() not in {"NG1", "NG2"}:
            return READINESS_NEEDS_CONFIRM
        return READINESS_READY
    if activity_type == "diesel":
        if year is not None and year < HEATING_VALUE_READY_YEAR:
            return READINESS_UNSUPPORTED
        if str(process_use or "").strip() != "company_vehicle":
            return READINESS_NEEDS_CONFIRM
        return READINESS_READY
    if activity_type == "grid_electricity":
        if (
            year is not None
            and year >= HEATING_VALUE_READY_YEAR
            and str(process_use or "").strip() not in _ENTERPRISE_ELECTRICITY_USES
        ):
            return READINESS_NEEDS_CONFIRM
        return READINESS_READY
    return READINESS_UNSUPPORTED


def summarize_pre_analysis_readiness(accepted: pd.DataFrame) -> dict[str, int]:
    """Count ready / needs-confirm / unsupported accepted activities."""
    summary = {
        READINESS_READY: 0,
        READINESS_NEEDS_CONFIRM: 0,
        READINESS_UNSUPPORTED: 0,
    }
    if accepted is None or getattr(accepted, "empty", True):
        return summary
    for _, row in accepted.iterrows():
        bucket = classify_activity_analysis_readiness(
            activity_type=str(row.get("activity_type") or ""),
            fuel_subtype=str(row.get("fuel_subtype") or ""),
            process_use=str(row.get("process_use") or ""),
            activity_start=row.get("activity_start_date"),
            activity_end=row.get("activity_end_date"),
        )
        summary[bucket] = int(summary.get(bucket, 0)) + 1
    return summary


def context_confirmations_needed(
    uploaded: UploadedTable,
    mapping: ColumnMapping,
) -> dict[str, bool]:
    """Which mapping confirmations are still required for this upload."""
    needed = {"natural_gas": False, "diesel": False, "electricity": False}
    frame = uploaded.frame
    activity_col = mapping.activity_type_column
    if not activity_col or activity_col not in frame.columns:
        return needed
    value_map = mapping.activity_type_value_map or {}
    subtype_col = str(mapping.natural_gas_subtype_column or "").strip()
    file_ng = str(mapping.natural_gas_subtype or "").strip()
    file_diesel = str(mapping.diesel_context or "").strip()
    file_electricity = str(mapping.electricity_context or "").strip()
    for _, row in frame.iterrows():
        source = row.get(activity_col)
        source_text = str(source).strip() if source is not None else ""
        mapped = value_map.get(source_text) or suggest_activity_type(source_text)
        if mapped == "natural_gas":
            subtype = extract_natural_gas_subtype_from_text(source_text)
            if (
                not subtype
                and subtype_col
                and subtype_col in frame.columns
            ):
                subtype = extract_natural_gas_subtype_from_text(row.get(subtype_col))
            if not subtype and file_ng not in {"NG1", "NG2"}:
                needed["natural_gas"] = True
        elif mapped == "diesel":
            if (
                not extract_diesel_vehicle_context_from_text(source_text)
                and file_diesel != "company_vehicle"
            ):
                needed["diesel"] = True
        elif mapped == "grid_electricity":
            if file_electricity != "enterprise":
                needed["electricity"] = True
    return needed


def _transport_payer_for_record_type(record_type: str) -> str:
    if record_type == "transport_activity":
        return "unknown"
    return "not_applicable"


def _rejection(
    *,
    source_row: int,
    field_name: str,
    issue_code: str,
    issue_message: str,
    uploaded_value: Any,
) -> dict[str, Any]:
    return {
        "source_row": source_row,
        "field": field_name,
        "issue_code": issue_code,
        "issue_message": issue_message,
        "uploaded_value": (
            "" if uploaded_value is None else str(uploaded_value)
        ),
    }


def _as_naive_timestamp(value: Any) -> pd.Timestamp:
    """Normalize an explicit timestamp to timezone-naive pandas form.

    Preserves naive timestamps. Converts timezone-aware values to naive via
    tz_convert(None) without inventing a new instant.
    """
    timestamp = pd.Timestamp(value)
    if timestamp.tzinfo is not None:
        timestamp = timestamp.tz_convert(None)
    return timestamp


def build_source_document_row(
    uploaded: UploadedTable,
    metadata: IntakeMetadata,
) -> dict[str, Any]:
    """Build one canonical source-document row for the uploaded file."""
    if metadata.document_date is None:
        raise IntakeError(
            "DOCUMENT_DATE_REQUIRED",
            "Document date is required and must be confirmed; "
            "unknown must not become today's date.",
        )
    return {
        "source_document_id": source_document_id_from_hash(uploaded.sha256),
        "file_name": uploaded.file_name,
        "document_type": "other",
        "document_date": datetime(
            metadata.document_date.year,
            metadata.document_date.month,
            metadata.document_date.day,
        ),
        "issuer": pd.NA,
        "data_origin": "company_provided",
        "is_synthetic": False,
        "source_path": pd.NA,
        "sha256": uploaded.sha256,
        "ingested_at": _as_naive_timestamp(metadata.ingested_at),
        "ingestion_run_id": metadata.intake_run_id,
        "notes": SOURCE_DOCUMENT_NOTES,
    }


def _units_compatible(activity_type: str, unit: str) -> bool:
    allowed = ACTIVITY_TYPE_UNIT_COMPATIBILITY.get(activity_type)
    if allowed is None:
        return unit in SUPPORTED_UNITS
    return unit in allowed


def build_and_validate_intake(
    uploaded: UploadedTable,
    mapping: ColumnMapping,
    metadata: IntakeMetadata,
) -> IntakeValidationResult:
    """Build canonical rows and validate without mutating the uploaded frame."""
    source_frame = uploaded.frame.copy()
    rejected: list[dict[str, Any]] = []
    accepted: list[dict[str, Any]] = []

    required_columns = {
        "activity_type": mapping.activity_type_column,
        "activity_value": mapping.activity_value_column,
        "unit": mapping.unit_column,
    }
    for field_name, column in required_columns.items():
        if not column or column not in source_frame.columns:
            raise IntakeError(
                ISSUE_MISSING_REQUIRED_MAPPING,
                f"Missing required mapping for {field_name}.",
            )
        if is_reference_only_column(column):
            raise IntakeError(
                ISSUE_MISSING_REQUIRED_MAPPING,
                f"Column {column!r} is reference-only and cannot map to {field_name}.",
            )

    if mapping.use_year_month:
        if not mapping.year_month_column or mapping.year_month_column not in (
            source_frame.columns
        ):
            raise IntakeError(
                ISSUE_MISSING_REQUIRED_MAPPING,
                "Missing required year-month column mapping.",
            )
        if not mapping.year_month_confirmed:
            raise IntakeError(
                ISSUE_YEAR_MONTH_NOT_CONFIRMED,
                "Year-month transformation requires explicit user confirmation.",
            )
    elif mapping.use_file_dates:
        if (
            not mapping.start_date_column
            or mapping.start_date_column not in source_frame.columns
            or not mapping.end_date_column
            or mapping.end_date_column not in source_frame.columns
        ):
            raise IntakeError(
                ISSUE_MISSING_REQUIRED_MAPPING,
                "Missing required date column mapping.",
            )
    elif mapping.period_start is None or mapping.period_end is None:
        raise IntakeError(
            ISSUE_MISSING_REQUIRED_MAPPING,
            "Reporting period start and end dates are required.",
        )

    source_doc = build_source_document_row(uploaded, metadata)
    try:
        validate_source_documents(pd.DataFrame([source_doc]))
    except (SchemaError, SchemaErrors) as exc:
        raise IntakeError(
            ISSUE_SCHEMA_VALIDATION_FAILED,
            f"Source document failed schema validation: {exc}",
        ) from exc

    # Excel-style row numbers: header_row_index is 0-based; data starts next.
    data_row_base = uploaded.header_row_index + 2

    for offset, (_, row) in enumerate(source_frame.iterrows()):
        source_row = offset + data_row_base

        raw_activity = row.get(mapping.activity_type_column)
        raw_value = row.get(mapping.activity_value_column)
        raw_unit = row.get(mapping.unit_column)

        # Skip completely empty rows.
        if (
            (raw_activity is None or str(raw_activity).strip() == "")
            and (raw_value is None or str(raw_value).strip() == "")
            and (raw_unit is None or str(raw_unit).strip() == "")
        ):
            continue

        activity_key = str(raw_activity).strip() if raw_activity is not None else ""
        mapped_activity = mapping.activity_type_value_map.get(activity_key, "")
        if not mapped_activity or mapped_activity not in ACTIVITY_TYPES:
            rejected.append(
                _rejection(
                    source_row=source_row,
                    field_name="activity_type",
                    issue_code=ISSUE_UNMAPPED_ACTIVITY_TYPE,
                    issue_message=(
                        "Activity type is not mapped to a supported value."
                    ),
                    uploaded_value=raw_activity,
                )
            )
            continue

        unit_key = str(raw_unit).strip() if raw_unit is not None else ""
        mapped_unit = mapping.unit_value_map.get(unit_key, "")
        if not mapped_unit or mapped_unit not in SUPPORTED_UNITS:
            rejected.append(
                _rejection(
                    source_row=source_row,
                    field_name="unit",
                    issue_code=ISSUE_UNMAPPED_UNIT,
                    issue_message="Unit is not mapped to a supported value.",
                    uploaded_value=raw_unit,
                )
            )
            continue

        if not _units_compatible(mapped_activity, mapped_unit):
            rejected.append(
                _rejection(
                    source_row=source_row,
                    field_name="unit",
                    issue_code=ISSUE_ACTIVITY_UNIT_MISMATCH,
                    issue_message=(
                        f"Unit {mapped_unit} is not compatible with "
                        f"activity type {mapped_activity}."
                    ),
                    uploaded_value=raw_unit,
                )
            )
            continue

        try:
            activity_value = _parse_activity_value(raw_value)
        except ValueError as exc:
            rejected.append(
                _rejection(
                    source_row=source_row,
                    field_name="activity_value",
                    issue_code=ISSUE_INVALID_ACTIVITY_VALUE,
                    issue_message=str(exc),
                    uploaded_value=raw_value,
                )
            )
            continue

        try:
            if mapping.use_year_month:
                start_date, end_date = parse_year_month_period(
                    row.get(mapping.year_month_column)
                )
                start_dt = datetime(
                    start_date.year, start_date.month, start_date.day
                )
                end_dt = datetime(end_date.year, end_date.month, end_date.day)
            elif mapping.use_file_dates:
                start_dt = _parse_date(
                    row.get(mapping.start_date_column),
                    field_name="activity_start_date",
                )
                end_dt = _parse_date(
                    row.get(mapping.end_date_column),
                    field_name="activity_end_date",
                )
            else:
                assert mapping.period_start is not None
                assert mapping.period_end is not None
                start_dt = datetime(
                    mapping.period_start.year,
                    mapping.period_start.month,
                    mapping.period_start.day,
                )
                end_dt = datetime(
                    mapping.period_end.year,
                    mapping.period_end.month,
                    mapping.period_end.day,
                )
            if end_dt < start_dt:
                raise ValueError("end date must be on or after start date")
        except (ValueError, TypeError) as exc:
            rejected.append(
                _rejection(
                    source_row=source_row,
                    field_name="activity_start_date",
                    issue_code=ISSUE_INVALID_DATE,
                    issue_message=str(exc),
                    uploaded_value=(
                        row.get(mapping.year_month_column)
                        if mapping.use_year_month
                        else (
                            row.get(mapping.start_date_column)
                            if mapping.use_file_dates
                            else mapping.period_start
                        )
                    ),
                )
            )
            continue

        if mapping.site_column and mapping.site_column in source_frame.columns:
            site_raw = row.get(mapping.site_column)
            site_id = (
                str(site_raw).strip()
                if site_raw is not None and str(site_raw).strip()
                else (metadata.site_id.strip() or "UNKNOWN")
            )
        else:
            site_id = metadata.site_id.strip() or "UNKNOWN"

        record_type = record_type_for_activity(mapped_activity)
        process_use = _resolve_process_use(
            mapped_activity=mapped_activity,
            mapping=mapping,
            activity_source=activity_key,
        )
        ownership = _inventory_ownership(mapped_activity, process_use)
        org_boundary = _inventory_org_boundary(mapped_activity, process_use)
        transport_payer = _transport_payer_for_record_type(record_type)
        needs_review = True  # conservative unknowns remain

        notes: Any = pd.NA
        if mapped_activity == "other" or record_type == "other":
            notes = "User-mapped activity classified as other."

        fuel_subtype = "not_applicable"
        if mapped_activity == "natural_gas":
            fuel_subtype = _resolve_natural_gas_subtype(
                row=row,
                mapping=mapping,
                activity_source=activity_key,
            )

        # Intentionally ignore uploaded emission-factor / emission-result columns.
        # Those remain source/reference only; calculation uses the controlled registry.
        activity_row = {
            "record_id": activity_record_id(uploaded.sha256, source_row),
            "source_document_id": source_doc["source_document_id"],
            "source_locator": source_locator(
                sheet_name=uploaded.sheet_name,
                source_row=source_row,
            ),
            "record_type": record_type,
            "activity_start_date": start_dt,
            "activity_end_date": end_dt,
            "site_id": site_id,
            "production_process_id": pd.NA,
            "product_id": pd.NA,
            "activity_type": mapped_activity,
            "process_use": process_use,
            "fuel_subtype": fuel_subtype,
            "activity_value": float(activity_value),
            "unit": mapped_unit,
            "transport_payer": transport_payer,
            "ownership_control": ownership,
            "organizational_boundary_status": org_boundary,
            "cbam_process_boundary_status": "unknown",
            "measurement_method": "unknown",
            "data_quality_tier": metadata.data_quality_tier,
            "human_review_status": (
                "needs_review" if needs_review else "not_required"
            ),
            "notes": notes,
        }

        try:
            validate_activity_records(pd.DataFrame([activity_row]))
        except (SchemaError, SchemaErrors) as exc:
            rejected.append(
                _rejection(
                    source_row=source_row,
                    field_name="record",
                    issue_code=ISSUE_SCHEMA_VALIDATION_FAILED,
                    issue_message=str(exc),
                    uploaded_value=activity_key,
                )
            )
            continue

        accepted.append(activity_row)

    accepted_df = pd.DataFrame(accepted)
    rejected_df = pd.DataFrame(rejected, columns=list(REJECTION_COLUMNS))
    source_docs_df = pd.DataFrame([source_doc])
    total = int(len(accepted) + len(rejected))
    duplicate_groups = tuple(
        find_potential_duplicate_groups(
            accepted_df,
            file_hash=uploaded.sha256,
        )
    )
    return IntakeValidationResult(
        source_documents=source_docs_df,
        accepted_activities=accepted_df,
        rejected_rows=rejected_df,
        accepted_count=int(len(accepted)),
        rejected_count=int(len(rejected)),
        total_count=total,
        file_hash=uploaded.sha256,
        file_name=uploaded.file_name,
        potential_duplicate_groups=duplicate_groups,
    )


def default_value_maps(
    uploaded: UploadedTable,
    mapping: ColumnMapping,
) -> tuple[dict[str, str], dict[str, str]]:
    """Build suggested activity-type and unit maps for distinct values."""
    activity_map: dict[str, str] = {}
    for value in distinct_values(uploaded.frame, mapping.activity_type_column):
        activity_map[value] = suggest_activity_type(value)
    unit_map: dict[str, str] = {}
    for value in distinct_values(uploaded.frame, mapping.unit_column):
        unit_map[value] = suggest_unit(value)
    return activity_map, unit_map
