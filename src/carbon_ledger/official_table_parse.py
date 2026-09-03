"""Deterministic parsers for official MOENV ODS/XLSX emission-factor tables.

Parsers identify rows by sheet name, headers, fuel/gas labels, and units.
They do not guess from a lone Excel row number when the table layout changes.
"""

from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation
from typing import TYPE_CHECKING, Any
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile

if TYPE_CHECKING:
    from carbon_ledger.reference_sync import ParseResult

TABLE_NS = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
TEXT_NS = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"

PARSER_FUEL_ODS = "tw_moenv_general_emission_factors_ods_v1"
PARSER_GWP_ODS = "tw_moenv_gwp_ods_v1"

SHEET_NOTES = "說明"
SHEET_STATIONARY = "附表一_固定燃燒排放源排放係數"
SHEET_MOBILE = "附表一_移動燃燒排放源排放係數"
SHEET_GWP = "附表四"

REF_TYPE_FUEL_EF = "fuel_emission_factor"
REF_TYPE_GWP = "gwp_reference"
REF_TYPE_STEEL = "purchased_steel_average_data"

ASSESSMENT_AR5 = "IPCC AR5 100-year GWP"

_ROC_ANNOUNCEMENT_RE = re.compile(
    r"中華民國\s*([0-9]{2,3})\s*年\s*([0-9]{1,2})\s*月\s*([0-9]{1,2})\s*日"
)
_FIFTH_ASSESSMENT_RE = re.compile(
    r"Fifth Assessment Report|AR5|第五次評估",
    re.IGNORECASE,
)

REQUIRED_GWP_GASES: tuple[tuple[str, str, str], ...] = (
    ("二氧化碳", "CO2", "CO2"),
    ("甲烷（Methane）", "CH4", "CH4"),
    ("石化甲烷", "CH4", "CH4"),
    ("氧化亞氮", "N2O", "N2O"),
    ("HFC-32", "CH2F2", "HFC-32"),
    ("HFC-125", "CHF2CF3", "HFC-125"),
    ("HFC-134a", "CH2FCF3", "HFC-134a"),
    ("HFC-143a", "CH3CF3", "HFC-143a"),
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_decimal(raw: str) -> Decimal | None:
    token = _text(raw).replace(",", "").replace(" ", "")
    if not token:
        return None
    try:
        number = Decimal(token)
    except (InvalidOperation, ValueError):
        return None
    if not number.is_finite() or number <= 0:
        return None
    return number


def _roc_to_gregorian_date(year: str, month: str, day: str) -> str:
    try:
        gregorian = int(year) + 1911
        return f"{gregorian:04d}-{int(month):02d}-{int(day):02d}"
    except ValueError:
        return ""


def load_ods_tables(content: bytes) -> dict[str, list[list[str]]] | None:
    """Return sheet name → rows of cell text. None when the bytes are not ODS."""
    try:
        archive = ZipFile(io.BytesIO(content))
    except BadZipFile:
        return None
    try:
        xml = archive.read("content.xml")
    except KeyError:
        return None
    root = ET.fromstring(xml)
    tables: dict[str, list[list[str]]] = {}
    for table in root.findall(f".//{{{TABLE_NS}}}table"):
        name = _text(table.get(f"{{{TABLE_NS}}}name"))
        if not name:
            continue
        rows: list[list[str]] = []
        for row in table.findall(f"{{{TABLE_NS}}}table-row"):
            cells: list[str] = []
            for cell in row.findall(f"{{{TABLE_NS}}}table-cell"):
                repeat = int(cell.get(f"{{{TABLE_NS}}}number-columns-repeated", "1"))
                parts = [
                    "".join(paragraph.itertext())
                    for paragraph in cell.findall(f".//{{{TEXT_NS}}}p")
                ]
                text = " ".join(part.strip() for part in parts if part.strip())
                cells.extend([text] * min(max(repeat, 1), 24))
            rows.append(cells)
        tables[name] = rows
    return tables or None


def load_xlsx_tables(content: bytes) -> dict[str, list[list[str]]] | None:
    """Return sheet name → rows for an official XLSX workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return None
    tables: dict[str, list[list[str]]] = {}
    try:
        for sheet in workbook.worksheets:
            rows: list[list[str]] = []
            for raw in sheet.iter_rows(values_only=True):
                rows.append([_text(cell) for cell in raw])
            tables[_text(sheet.title)] = rows
    finally:
        workbook.close()
    return tables or None


def load_official_tables(content: bytes) -> dict[str, list[list[str]]] | None:
    tables = load_ods_tables(content)
    if tables is not None:
        return tables
    return load_xlsx_tables(content)


def _joined_row(row: list[str]) -> str:
    return " ".join(cell for cell in row if cell)


def _extract_publication_date(notes_rows: list[list[str]]) -> str:
    """Return the unique ROC announcement date, or blank when it is not unique."""
    blob = "\n".join(_joined_row(row) for row in notes_rows)
    matches = list(_ROC_ANNOUNCEMENT_RE.finditer(blob))
    if len(matches) != 1:
        return ""
    match = matches[0]
    return _roc_to_gregorian_date(match.group(1), match.group(2), match.group(3))


def _factor_year_from_publication(publication_date: str) -> str:
    """Official document/version year from a parsed announcement date."""
    year = _text(publication_date)[:4]
    if len(year) == 4 and year.isdigit():
        return year
    return ""


def _extract_gwp_assessment_basis(tables: dict[str, list[list[str]]]) -> str:
    notes = tables.get(SHEET_NOTES) or []
    gwp_rows = tables.get(SHEET_GWP) or []
    blob = "\n".join(
        _joined_row(row) for row in notes + gwp_rows
    )
    if _FIFTH_ASSESSMENT_RE.search(blob):
        return ASSESSMENT_AR5
    return ""


def _find_fuel_table_header(rows: list[list[str]]) -> int | None:
    """Locate the fuel-factor table header by labels, not a fixed row number."""
    for index, row in enumerate(rows):
        first = _text(row[0]) if row else ""
        blob = _joined_row(row)
        if (
            first == "燃料"
            and "英譯名稱" in blob
            and "燃料單位熱值之排放係數" in blob
        ):
            return index
    return None


def _find_header_index(rows: list[list[str]], *tokens: str) -> int | None:
    for index, row in enumerate(rows):
        blob = _joined_row(row)
        if all(token in blob for token in tokens):
            return index
    return None


def _parse_result_type():
    from carbon_ledger.reference_sync import ParseResult

    return ParseResult


def _review(parser_type: str, reason: str):
    from carbon_ledger.reference_sync import LIFECYCLE_NEEDS_PARSER_REVIEW

    return _parse_result_type()(
        parser_type=parser_type,
        status=LIFECYCLE_NEEDS_PARSER_REVIEW,
        reason=reason,
    )


def _kg_per_tj_data_columns(
    header_row: list[str],
    unit_row: list[str],
    gas_row: list[str],
) -> dict[str, int]:
    """Map CO2/CH4/N2O to data-row columns using headers, not a fixed Excel row."""
    unit_blob = _joined_row(unit_row)
    if "kg/TJ" not in unit_blob and "公斤/兆焦耳" not in unit_blob:
        return {}
    gases = [_text(cell).upper() for cell in gas_row]
    label_start = None
    for index in range(len(gases) - 2):
        if gases[index : index + 3] == ["CO2", "CH4", "N2O"]:
            label_start = index
            break
    if label_start is None:
        return {}
    data_start = label_start
    if _text(header_row[0] if header_row else "") == "燃料" and label_start < 2:
        data_start = 2
    return {"CO2": data_start, "CH4": data_start + 1, "N2O": data_start + 2}


def _match_fuel_row(
    rows: list[list[str]],
    *,
    chinese: str,
    english_contains: tuple[str, ...],
    start: int,
) -> list[str] | None:
    matches: list[list[str]] = []
    for row in rows[start:]:
        if not row:
            continue
        first = _text(row[0])
        if first.startswith("註"):
            break
        if first != chinese:
            continue
        english = _text(row[1]) if len(row) > 1 else ""
        if any(token.lower() in english.lower() for token in english_contains):
            matches.append(row)
    if len(matches) != 1:
        return None
    return matches[0]


def parse_moenv_ods_fuel_emission_factors(content: bytes) -> ParseResult:
    """Parse stationary natural-gas and mobile diesel kg/TJ factors from ODS/XLSX."""
    tables = load_official_tables(content)
    if tables is None:
        return _review(
            PARSER_FUEL_ODS,
            "Bytes are not a readable ODS or XLSX workbook.",
        )
    missing = [
        name
        for name in (SHEET_NOTES, SHEET_STATIONARY, SHEET_MOBILE)
        if name not in tables
    ]
    if missing:
        return _review(
            PARSER_FUEL_ODS,
            "Official fuel-factor sheets missing or renamed: " + ", ".join(missing),
        )
    notes_blob = "\n".join(_joined_row(row) for row in tables[SHEET_NOTES])
    if "溫室氣體排放係數" not in notes_blob:
        return _review(
            PARSER_FUEL_ODS,
            "Notes sheet does not identify 溫室氣體排放係數; format may have changed.",
        )
    publication_date = _extract_publication_date(tables[SHEET_NOTES])
    if not publication_date:
        return _review(
            PARSER_FUEL_ODS,
            (
                "Announcement date on the notes sheet could not be parsed; "
                "manual_review_required."
            ),
        )
    factor_year = _factor_year_from_publication(publication_date)
    if not factor_year:
        return _review(
            PARSER_FUEL_ODS,
            (
                "Official document year could not be taken from the "
                "announcement date; manual_review_required."
            ),
        )

    stationary = tables[SHEET_STATIONARY]
    fuel_header = _find_fuel_table_header(stationary)
    if fuel_header is None:
        return _review(
            PARSER_FUEL_ODS,
            "Stationary sheet is missing 燃料 / 燃料單位熱值之排放係數 headers.",
        )
    if fuel_header + 2 >= len(stationary):
        return _review(
            PARSER_FUEL_ODS,
            "Stationary sheet header/unit/gas rows are incomplete.",
        )
    unit_row = stationary[fuel_header + 1]
    gas_row = stationary[fuel_header + 2]
    gas_cols = _kg_per_tj_data_columns(stationary[fuel_header], unit_row, gas_row)
    if set(gas_cols) != {"CO2", "CH4", "N2O"}:
        return _review(
            PARSER_FUEL_ODS,
            "Stationary kg/TJ CO2/CH4/N2O columns could not be identified.",
        )
    ng_row = _match_fuel_row(
        stationary,
        chinese="天然氣",
        english_contains=("Natural Gas",),
        start=fuel_header + 3,
    )
    if ng_row is None or "Liquids" in _text(ng_row[1] if len(ng_row) > 1 else ""):
        return _review(
            PARSER_FUEL_ODS,
            "Stationary Natural Gas (天然氣) kg/TJ row was not uniquely identified.",
        )
    ng_values = {
        gas: _parse_decimal(ng_row[index] if index < len(ng_row) else "")
        for gas, index in gas_cols.items()
    }
    if any(value is None for value in ng_values.values()):
        return _review(
            PARSER_FUEL_ODS,
            "Stationary Natural Gas kg/TJ values are missing or not finite.",
        )

    mobile = tables[SHEET_MOBILE]
    co2_header = _find_fuel_table_header(mobile)
    if co2_header is None:
        return _review(PARSER_FUEL_ODS, "Mobile sheet is missing 燃料 headers.")
    mobile_unit = mobile[co2_header + 1] if co2_header + 1 < len(mobile) else []
    mobile_gas = mobile[co2_header + 2] if co2_header + 2 < len(mobile) else []
    if "CO2" not in _joined_row(mobile_gas):
        return _review(PARSER_FUEL_ODS, "Mobile CO2 header row was not found.")
    if "kg/TJ" not in _joined_row(mobile_unit) and "公斤/兆焦耳" not in _joined_row(
        mobile_unit
    ):
        return _review(PARSER_FUEL_ODS, "Mobile kg/TJ unit header was not found.")
    diesel_co2_row = _match_fuel_row(
        mobile,
        chinese="柴油",
        english_contains=("Gas/ Diesel", "Gas/Diesel"),
        start=co2_header + 3,
    )
    if diesel_co2_row is None:
        return _review(
            PARSER_FUEL_ODS,
            "Mobile diesel CO2 kg/TJ row was not uniquely identified.",
        )
    diesel_co2 = _parse_decimal(diesel_co2_row[2] if len(diesel_co2_row) > 2 else "")
    if diesel_co2 is None:
        return _review(PARSER_FUEL_ODS, "Mobile diesel CO2 kg/TJ value is not finite.")

    ch4_header = None
    for index, row in enumerate(mobile):
        blob = _joined_row(row)
        if "CH4" in blob and "N2O" in blob and "CO2" not in blob.split("CH4")[0]:
            if index > co2_header + 3:
                ch4_header = index
                break
    if ch4_header is None:
        return _review(PARSER_FUEL_ODS, "Mobile CH4/N2O header row was not found.")
    diesel_ch4_row = _match_fuel_row(
        mobile,
        chinese="柴油",
        english_contains=("Diesel Oil",),
        start=ch4_header + 1,
    )
    if diesel_ch4_row is None:
        return _review(
            PARSER_FUEL_ODS,
            "Mobile diesel CH4/N2O kg/TJ row was not uniquely identified.",
        )
    diesel_ch4 = _parse_decimal(diesel_ch4_row[2] if len(diesel_ch4_row) > 2 else "")
    diesel_n2o = _parse_decimal(diesel_ch4_row[3] if len(diesel_ch4_row) > 3 else "")
    if diesel_ch4 is None or diesel_n2o is None:
        return _review(
            PARSER_FUEL_ODS,
            "Mobile diesel CH4/N2O kg/TJ values are missing or not finite.",
        )

    records: list[dict[str, Any]] = []
    for gas, value in ng_values.items():
        assert value is not None
        records.append(
            _fuel_record(
                activity_type="natural_gas",
                combustion_context="stationary_combustion",
                gas=gas,
                value=value,
                publication_date=publication_date,
                source_locator=(
                    f"ODS sheet {SHEET_STATIONARY}; fuel=天然氣 / Natural Gas; "
                    f"gas={gas}; unit=kg/TJ"
                ),
                factor_year=factor_year,
            )
        )
    records.append(
        _fuel_record(
            activity_type="diesel",
            combustion_context="mobile_combustion",
            gas="CO2",
            value=diesel_co2,
            publication_date=publication_date,
            source_locator=(
                f"ODS sheet {SHEET_MOBILE}; fuel=柴油 / Gas/ Diesel; "
                "gas=CO2; unit=kg/TJ"
            ),
            factor_year=factor_year,
        )
    )
    records.append(
        _fuel_record(
            activity_type="diesel",
            combustion_context="mobile_combustion",
            gas="CH4",
            value=diesel_ch4,
            publication_date=publication_date,
            source_locator=(
                f"ODS sheet {SHEET_MOBILE}; fuel=柴油 / Gas / Diesel Oil; "
                "gas=CH4; unit=kg/TJ"
            ),
            factor_year=factor_year,
        )
    )
    records.append(
        _fuel_record(
            activity_type="diesel",
            combustion_context="mobile_combustion",
            gas="N2O",
            value=diesel_n2o,
            publication_date=publication_date,
            source_locator=(
                f"ODS sheet {SHEET_MOBILE}; fuel=柴油 / Gas / Diesel Oil; "
                "gas=N2O; unit=kg/TJ"
            ),
            factor_year=factor_year,
        )
    )
    from carbon_ledger.reference_sync import LIFECYCLE_PARSED

    return _parse_result_type()(
        parser_type=PARSER_FUEL_ODS,
        status=LIFECYCLE_PARSED,
        records=records,
        publication_date=publication_date,
        reason=(
            "Parsed official MOENV stationary natural-gas and mobile diesel "
            "kg/TJ factors. Applicability period was not stated, so "
            "valid_from/valid_to were left blank."
        ),
    )


def _fuel_record(
    *,
    activity_type: str,
    combustion_context: str,
    gas: str,
    value: Decimal,
    publication_date: str,
    source_locator: str,
    factor_year: str,
) -> dict[str, Any]:
    numerator = {"CO2": "kgCO2", "CH4": "kgCH4", "N2O": "kgN2O"}[gas]
    return {
        "reference_type": REF_TYPE_FUEL_EF,
        "candidate_type": REF_TYPE_FUEL_EF,
        "target_registry": "emission_factors",
        "factor_year": factor_year,
        "reporting_year": "",
        "factor_value": format(value, "f"),
        "numerator_unit": numerator,
        "denominator_unit": "TJ",
        "factor_unit": f"{numerator}/TJ",
        "geography": "TW_reference",
        "activity_type": activity_type,
        "combustion_context": combustion_context,
        "factor_context": combustion_context,
        "gas": gas,
        "factor_category": activity_type,
        "valid_from": "",
        "valid_to": "",
        "publication_date": publication_date,
        "source_locator": source_locator,
        "assessment_basis": "",
        "refrigerant": "",
    }


def parse_moenv_ods_gwp(content: bytes) -> ParseResult:
    """Parse AR5 100-year GWP values from official 附表四."""
    tables = load_official_tables(content)
    if tables is None:
        return _review(PARSER_GWP_ODS, "Bytes are not a readable ODS or XLSX workbook.")
    if SHEET_GWP not in tables or SHEET_NOTES not in tables:
        return _review(
            PARSER_GWP_ODS,
            "GWP sheet 附表四 or notes sheet is missing or renamed.",
        )
    assessment = _extract_gwp_assessment_basis(tables)
    if assessment != ASSESSMENT_AR5:
        return _review(
            PARSER_GWP_ODS,
            (
                "GWP assessment basis is not identifiably IPCC AR5; "
                "values were not guessed."
            ),
        )
    publication_date = _extract_publication_date(tables[SHEET_NOTES])
    if not publication_date:
        return _review(
            PARSER_GWP_ODS,
            (
                "Announcement date on the notes sheet could not be parsed; "
                "manual_review_required."
            ),
        )
    factor_year = _factor_year_from_publication(publication_date)
    if not factor_year:
        return _review(
            PARSER_GWP_ODS,
            (
                "Official document year could not be taken from the "
                "announcement date; manual_review_required."
            ),
        )
    rows = tables[SHEET_GWP]
    header = _find_header_index(rows, "溫暖化潛勢")
    if header is None:
        return _review(PARSER_GWP_ODS, "附表四 is missing the 溫暖化潛勢 header.")
    found: dict[str, dict[str, Any]] = {}
    for row in rows[header + 1 :]:
        if len(row) < 3:
            continue
        label, formula, raw_value = _text(row[0]), _text(row[1]), _text(row[2])
        if not label or label.startswith("註"):
            continue
        value = _parse_decimal(raw_value)
        if value is None:
            continue
        for token, expected_formula, canonical in REQUIRED_GWP_GASES:
            if token not in label:
                continue
            if expected_formula and formula and formula != expected_formula:
                continue
            key = canonical
            if token == "石化甲烷":
                key = "fossil_methane"
            if key in found:
                return _review(
                    PARSER_GWP_ODS,
                    f"Duplicate GWP row for {canonical}.",
                )
            context = "fuel_combustion"
            refrigerant = ""
            if canonical.startswith("HFC-"):
                context = "refrigerant_fugitive"
                refrigerant = canonical
            if key == "fossil_methane":
                context = "fossil_methane_process"
                canonical_gas = "CH4"
            else:
                canonical_gas = canonical
            found[key] = {
                "reference_type": REF_TYPE_GWP,
                "candidate_type": REF_TYPE_GWP,
                "target_registry": "gwp_values",
                "factor_year": factor_year,
                "reporting_year": "",
                "factor_value": format(value, "f"),
                "numerator_unit": "GWP",
                "denominator_unit": "",
                "factor_unit": "GWP",
                "geography": "TW",
                "activity_type": "gwp",
                "combustion_context": context,
                "factor_context": context,
                "gas": canonical_gas,
                "refrigerant": refrigerant,
                "assessment_basis": assessment,
                "factor_category": context,
                "valid_from": "",
                "valid_to": "",
                "publication_date": publication_date,
                "source_locator": (
                    f"ODS sheet {SHEET_GWP}; label={label}; formula={formula}"
                ),
            }
    required_keys = {
        "CO2",
        "CH4",
        "fossil_methane",
        "N2O",
        "HFC-32",
        "HFC-125",
        "HFC-134a",
        "HFC-143a",
    }
    missing = sorted(required_keys - set(found))
    if missing:
        return _review(
            PARSER_GWP_ODS,
            "Required GWP gases were not found by name/formula: " + ", ".join(missing),
        )
    from carbon_ledger.reference_sync import LIFECYCLE_PARSED

    return _parse_result_type()(
        parser_type=PARSER_GWP_ODS,
        status=LIFECYCLE_PARSED,
        records=list(found.values()),
        publication_date=publication_date,
        reason=(
            "Parsed official MOENV 附表四 AR5 GWP values by gas name and formula. "
            "Applicability period was not stated, so valid_from/valid_to were "
            "left blank."
        ),
    )


def steel_average_data_not_configured_result():
    """Steel secondary factors have no approved official table in v1."""
    from carbon_ledger.reference_sync import LIFECYCLE_NEEDS_PARSER_REVIEW

    return _parse_result_type()(
        parser_type="purchased_steel_average_data_v1",
        status=LIFECYCLE_NEEDS_PARSER_REVIEW,
        records=[],
        reason=(
            "No approved purchased-steel average-data factor is configured. "
            "V1 does not invent or auto-activate a generic steel coefficient."
        ),
    )
